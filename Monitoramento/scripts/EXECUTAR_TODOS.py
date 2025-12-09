from openpyxl import Workbook
from pathlib import Path 
import os
import sys
import pandas as pd
from dotenv import load_dotenv 



pasta_scripts = Path(__file__).resolve().parent


pasta_inputs = pasta_scripts.parent / "inputs"
pasta_saida = pasta_scripts.parent / "outputs"
pasta_raiz_projeto = pasta_scripts.parent.parent


sys.path.append(str(pasta_raiz_projeto))

# Carrega variáveis de ambiente
load_dotenv()


try:
    from lib_validacao import export_query_to_csv
except ImportError:
    # Fallback caso o Python se perca nos caminhos relativos
    sys.path.append(str(pasta_scripts))
    from lib_validacao import export_query_to_csv

# --- INICIALIZAÇÃO DE WORKBOOKS ---
belem_wb = Workbook()
expansao_wb = Workbook()
grs_wb = Workbook()
expansao_ms_wb = Workbook()

# Remove aba padrão
for wb in [belem_wb, expansao_wb, grs_wb, expansao_ms_wb]:
    wb.remove(wb.active)

# Disponibiliza variáveis globais para os scripts executados via exec()

globals().update({
    "belem_wb": belem_wb,
    "expansao_wb": expansao_wb,
    "grs_wb": grs_wb,
    "expansao_ms_wb": expansao_ms_wb,
    "pasta_inputs": pasta_inputs, 
    "pasta_saida": pasta_saida
})

# --- ETAPA 1: ATUALIZAÇÃO DOS DADOS (BANCO -> CSV) ---

print("\n=== ETAPA 1: Atualizando Bases de Dados ===")

db_vars = ["DB_NAME", "DB_USER", "DB_PASSWORD", "DB_HOST"]
tem_credenciais = all(os.getenv(var) for var in db_vars)

mapa_queries = [
    ("form1.sql", "form1.csv"),
    ("form2.sql", "form2.csv"),
    ("form3.sql", "form3.csv"),
    ("form4.sql", "form4.csv")
]

if tem_credenciais:
    print("Credenciais encontradas. Iniciando extração...")
    try:
        for sql_file, csv_file in mapa_queries:
            print(f"Processando: {sql_file} -> {csv_file}...")
            export_query_to_csv(sql_file, csv_file, pasta_inputs)
        print(">>> Bases atualizadas com sucesso.")
    except Exception as e:
        print(f"[ERRO CRÍTICO] Falha na conexão ou extração: {e}")
else:
    print("[AVISO] Sem credenciais. Usando CSVs locais.")

# --- ETAPA 2: EXECUÇÃO DOS SCRIPTS ---

print("\n=== ETAPA 2: Gerando Planilhas de Monitoramento ===")

# Lista dos scripts que serão rodados
scripts_para_rodar = [
    "script_form1.py",
    "script_form2.py",
    "script_form3.py",
    "script_form4.py"
]

# Verifica CSVs
for _, csv_file in mapa_queries:
    if not (pasta_inputs / csv_file).exists():
        print(f"[ALERTA] Arquivo {csv_file} não encontrado em: {pasta_inputs}")

try:
    for script_nome in scripts_para_rodar:
        caminho_script_filho = pasta_scripts / script_nome
        
        if caminho_script_filho.exists():
            print(f"Executando {script_nome}...")
            # O exec lê o arquivo usando o caminho completo (pathlib)
            with open(caminho_script_filho, 'r', encoding='utf-8') as f:
                exec(f.read(), globals()) # Passa as globais explicitamente
            print(f"{script_nome}: OK")
        else:
            print(f"[ERRO] Script não encontrado: {caminho_script_filho}")

except Exception as e:
    print(f"[ERRO] Falha durante a execução dos scripts: {e}")


# --- ETAPA 3: VALIDAÇÃO ---

print("\n=== ETAPA 3: Validação ===")

if tem_credenciais:
    try:
        val_script = pasta_scripts / "script_validacao.py"
        with open(val_script, 'r', encoding='utf-8') as f:
            exec(f.read(), globals())
        print(">>> Validação finalizada.")
    except Exception as e:
        print(f"[ERRO] Falha na validação: {e}")
else:
    print("[PULADO] Validação requer credenciais.")


# --- ETAPA 4: SALVAMENTO ---

print("\n=== ETAPA 4: Salvando Arquivos ===")

mapa_saida = {
    belem_wb: "Belém",
    expansao_wb: "Expansão",
    grs_wb: "GRS",
    expansao_ms_wb: "Expansão MS"
}

for wb, nome_pasta in mapa_saida.items():
    # Cria caminho: outputs/NomePasta
    caminho_final_pasta = pasta_saida / nome_pasta
    caminho_final_pasta.mkdir(parents=True, exist_ok=True)
    
    caminho_arquivo = caminho_final_pasta / "0 - Monitoramento Form 1, 2 e 3.xlsx"
    
    if len(wb.sheetnames) > 0:
        wb.save(caminho_arquivo)
        print(f"Salvo em: {caminho_arquivo}")
    else:
        print(f"[AVISO] {nome_pasta} vazio. Não salvo.")

print("\nProcesso Finalizado.")