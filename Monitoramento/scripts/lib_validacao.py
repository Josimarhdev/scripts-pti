import os
import pandas as pd
import psycopg2
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from pathlib import Path


# (db.py)
def get_db_connection():
    return psycopg2.connect(
        dbname=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        host=os.getenv("DB_HOST"),
        port=os.getenv("DB_PORT")
    )

def export_query_to_csv(query_file_name, output_csv_name, pasta_inputs):
    """
    Lê o arquivo SQL da pasta inputs, executa e salva o CSV na mesma pasta inputs.
    """
    # Define os caminhos
    caminho_sql = pasta_inputs / query_file_name
    caminho_csv = pasta_inputs / output_csv_name

    if not caminho_sql.exists():
        print(f"[ERRO] Arquivo SQL não encontrado: {caminho_sql}")
        return

    try:
        conn = get_db_connection()
        query = caminho_sql.read_text(encoding='utf-8')
        
        df = pd.read_sql_query(query, conn)
        
        # Salva na pasta INPUTS
        df.to_csv(caminho_csv, index=False, encoding="utf-8-sig")
        print(f"[OK] Dados extraídos e salvos em: {caminho_csv.name}")
        
        conn.close()
    except Exception as e:
        print(f"[ERRO] Falha ao conectar ou salvar {output_csv_name}: {e}")


# (saver.py)
def __adjust_column_size(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                try:
                    max_length = max(max_length, len(str(cell.value)) + 5)
                except:
                    pass
        ws.column_dimensions[column_letter].width = max_length + 2

def __adjust_borders(ws):
    border_style = Border(
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000'),
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000')
    )
    for row in ws.iter_rows(max_col=ws.max_column):
        for cell in row:
            cell.border = border_style

def __align_rows(ws):
    for row in ws.iter_rows(max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

def __compare_cells(ws, row, col1, col2):
    v1 = ws[f"{col1}{row}"].value
    v2 = ws[f"{col2}{row}"].value
    return v1 == v2

def __paint(ws):
    # Cabeçalho
    for cell in ws["1:1"]:
        cell.font = Font(color="FFFFFF")
        cell.fill = PatternFill(start_color="244062", end_color="244062", fill_type="solid")

    for cell in ws["E:E"]:
        if cell.row == 1: continue
        
        row = cell.row
        # Regras de pintura condicionais
        if not __compare_cells(ws, row, "F", "G"):
            ws[f"F{row}"].fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")

        if not __compare_cells(ws, row, "H", "I"):
            ws[f"H{row}"].fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")

        if not __compare_cells(ws, row, "J", "K"):
            ws[f"J{row}"].fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")

        val_l = ws[f"L{row}"].value
        if val_l != 0 and __compare_cells(ws, row, "L", "M"):
             ws[f"L{row}"].fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")

    # Notas Fiscais
    for cell in ws["N:N"]:
        if cell.row == 1: continue
        if str(cell.value).lower() == "não":
            cell.fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")
    
    # Percentual Despesas
    for col in ["Q", "R"]:
        for cell in ws[f"{col}:{col}"]:
            if cell.row == 1: continue
            try:
                if float(cell.value) >= 60:
                    cell.fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")
            except:
                pass

    # Cores Regionais
    colors = {
        "Gabriel": "a9c5e6", "Bianca": "ffff99", "Valquiria": "b2ffff",
        "Larissa": "f1e0c6", "Luana": "ffccff", "Paranavaí": "9b59b6",
        "Maringá": "ffccff", "Francisco Beltrão": "b2ffff",
        "Londrina": "a9c5e6", "Guarapuava": "f1e0c6", "Curitiba": "ffff99"
    }
    for cell in ws["A:A"]:
        if cell.row == 1: continue
        if cell.value in colors:
            cell.fill = PatternFill(start_color=colors[cell.value], end_color=colors[cell.value], fill_type="solid")

def __add_regionals(ws):
    grs = {
        "Altônia": "Gabriel", "Anahy": "Bianca", "Assis Chateaubriand": "Gabriel",
        "Boa Vista da Aparecida": "Bianca", "Braganey": "Bianca", "Brasilândia do Sul": "Gabriel",
        "Cafelândia": "Valquiria", "Campo Bonito": "Bianca", "Capitão Leônidas Marques": "Bianca",
        "Cascavel": "Valquiria", "Catanduvas": "Bianca", "Céu Azul": "Luana", "Corbélia": "Bianca",
        "Diamante D'Oeste": "Luana", "Diamante do Sul": "Bianca", "Entre Rios do Oeste": "Luana",
        "Formosa do Oeste": "Valquiria", "Foz do Iguaçu": "Larissa",
        "Francisco Alves": "Gabriel", "Guaíra": "Gabriel",
        "Guaraniaçu": "Bianca", "Ibema": "Bianca", "Iguatu": "Bianca",
        "Iracema do Oeste": "Valquiria", "Itaipulândia": "Luana", "Jesuítas": "Valquiria",
        "Lindoeste": "Bianca", "Marechal Cândido Rondon": "Luana",
        "Maripá": "Gabriel", "Matelândia": "Larissa", "Medianeira": "Larissa",
        "Mercedes": "Gabriel", "Missal": "Luana", "Mundo Novo": "Gabriel",
        "Nova Aurora": "Valquiria", "Nova Santa Rosa": "Gabriel", "Ouro Verde do Oeste": "Gabriel",
        "Palotina": "Gabriel", "Pato Bragado": "Luana", "Quatro Pontes": "Gabriel",
        "Ramilândia": "Luana", "Santa Helena": "Luana", "Santa Lúcia": "Bianca",
        "Santa Tereza do Oeste": "Luana", "Santa Terezinha de Itaipu": "Larissa",
        "São José das Palmeiras": "Luana", "São Miguel do Iguaçu": "Larissa", "São Pedro do Iguaçu": "Luana",
        "Serranópolis do Iguaçu": "Larissa", "Terra Roxa": "Gabriel", "Toledo": "Gabriel",
        "Três Barras do Paraná": "Bianca", "Tupãssi": "Valquiria", "Ubiratã": "Valquiria",
        "Vera Cruz do Oeste": "Luana"
    }

    expansao = {
        " Alto Paraná": "Paranavaí", " Amaporã": "Paranavaí", "Apucarana": "Maringá",
        "Atalaia": "Paranavaí", "Barracão": "Francisco Beltrão", "Borrazópolis": "Maringá",
        "Cambará": "Londrina", "Campina da Lagoa": "Guarapuava",
        "Campo Largo": "Curitiba", "Campo Magro": "Curitiba", "Capanema": "Francisco Beltrão",
        "Cerro Azul": "Curitiba", "Colorado": "Paranavaí", "Coronel Vivida": "Laranjeiras do Sul",
        "Cruzeiro do Sul": "Paranavaí", "Enéas Marques": "Francisco Beltrão",
        "Francisco Beltrão": "Francisco Beltrão", "General Carneiro": "Laranjeiras do Sul",
        "Ibaiti": "Londrina", "Jaguapitã": "Londrina", "Jaguariaíva": "Londrina",
        "Jardim Alegre": "Maringá", "Kaloré": "Maringá", "Lapa": "Curitiba",
        "Laranjeiras do Sul": "Laranjeiras do Sul", "Mandaguaçu": "Maringá", "Mandaguari": "Maringá",
        "Marmeleiro": "Francisco Beltrão", "Mato Rico": "Guarapuava", "Nova Tebas": "Guarapuava",
        "Ortigueira": "Londrina", "Palmital": "Laranjeiras do Sul", "Paraíso do Norte": "Paranavaí",
        "Presidente Castelo Branco": "Maringá", "Quedas do Iguaçu": "Laranjeiras do Sul", "Realeza": "Francisco Beltrão",
        "Reserva do Iguaçu": "Laranjeiras do Sul", "Ribeirão Claro": "Londrina", "Rio Branco do Sul": "Curitiba",
        "Rio Negro": "Curitiba", "Salgado Filho": "Francisco Beltrão",
        "Santa Maria do Oeste": "Laranjeiras do Sul", "Santa Mariana": "Londrina", "Santa Mônica": "Paranavaí",
        "Santana do Itararé": "Londrina", "São João": "Laranjeiras do Sul", "São João do Caiuá": "Paranavaí",
        "São João do Ivaí": "Maringá", "São João do Triunfo": "Curitiba", "Tijucas do Sul": "Curitiba"
    }

    ws.insert_cols(1)
    ws["A1"].value = "regional"

    for row in ws["C:C"]:
        if row.row == 1: continue
        
        # strip() remove espaços extras que podem atrapalhar o match
        cidade = str(row.value).strip() if row.value else ""
        
        if cidade in expansao:
            ws[f"A{row.row}"] = expansao[cidade]
        elif cidade in grs:
            ws[f"A{row.row}"] = grs[cidade]
        # Tenta match direto também (caso a chave no dict tenha espaço)
        elif row.value in expansao:
             ws[f"A{row.row}"] = expansao[row.value]

def __delete_null(ws):
    rows_to_delete = []
    for cell in ws["E:E"]:
        if cell.row == 1: continue
        if cell.value is None:
            rows_to_delete.append(cell.row)
    
    for r in reversed(rows_to_delete):
        ws.delete_rows(r, 1)

def __add_filters(ws):
    ws.auto_filter.ref = ws.dimensions

def processar_e_salvar_excel(df, caminho_arquivo_saida):
    """
    Salva o DataFrame em Excel e aplica toda a formatação.
    """
    # 1. Cria o arquivo bruto
    df.to_excel(caminho_arquivo_saida, index=False)

    # 2. Reabre com Openpyxl para estilizar
    wb = load_workbook(caminho_arquivo_saida)
    ws = wb.active

    # Aplica as transformações
    __add_regionals(ws)
    __delete_null(ws)
    __add_filters(ws)
    __adjust_column_size(ws)
    __adjust_borders(ws)
    __align_rows(ws)
    __paint(ws)

    wb.save(caminho_arquivo_saida)
    print(f"[OK] Relatório formatado salvo em: {caminho_arquivo_saida}")