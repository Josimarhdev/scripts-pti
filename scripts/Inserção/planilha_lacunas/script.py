

import pandas as pd
from fuzzywuzzy import process, fuzz
import os
from openpyxl import load_workbook
import re

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

DEBUG_MODE = False
# Limiar de similaridade para fazer a correspondência 
MATCH_THRESHOLD = 70

#FUNÇÃO DE LIMPEZA DE TEXTO
def clean_text(text):

    if not isinstance(text, str):
        return ""
    text = re.sub(r'[()\[\],;/-]', ' ', text)
    text = re.sub(r'[^\w\s]', '', text)
    text = text.lower().strip()
    return " ".join(text.split())

# MAPA DE SINÔNIMOS
SYNONYM_MAP = {
    clean_text("Equipamentos de Proteção Individual"): "EPI",
    clean_text("Conserto e manutenção de equipamentos/veículo"): "Conserto e Manutenção de Equipamentos / Veículos",
    clean_text("Contratos"): "Prestação de Serviço",
    clean_text("PET Colorido (Embalagens de PET colorido, PET bagunça)"): "Colorido"

}

#CARREGAMENTO DOS ARQUIVOS
try:
    dados_path = os.path.join(SCRIPT_DIR,'inputs', 'dados.xlsx')
    template_path = os.path.join(SCRIPT_DIR,'inputs', 'template.xlsx')

    df_micro_dados = pd.read_excel(dados_path, sheet_name='Micro Dados - Filtrados')
    df_macro_dados = pd.read_excel(dados_path, sheet_name='Macro Dados - Filtrados')
    
    template_receitas_df = pd.read_excel(template_path, sheet_name='Receitas')
    template_despesas_df = pd.read_excel(template_path, sheet_name='Despesas', header=1)
    template_materiais_df = pd.read_excel(template_path, sheet_name='Materiais', header=1)
    template_materiais_df['Categoria'].ffill(inplace=True)

except FileNotFoundError as e:
    print(f"Erro: Arquivo não encontrado - {e}. Verifique se os arquivos 'dados.xlsx' e 'template.xlsx' estão na pasta 'inputs'.")
    exit()

#PREPARAÇÃO DOS DADOS
df_macro_dados['Data de referência'] = pd.to_datetime(df_macro_dados['Data de referência'], errors='coerce')
df_micro_dados['Data de referência'] = pd.to_datetime(df_micro_dados['Data de referência'], errors='coerce')
df_macro_dados.dropna(subset=['Município', 'UVR Número', 'Data de referência'], inplace=True)
df_macro_dados['UVR Número'] = df_macro_dados['UVR Número'].astype(int)
df_macro_dados['MesAno'] = df_macro_dados['Data de referência'].dt.strftime('%m-%Y')
df_micro_dados['MesAno'] = df_micro_dados['Data de referência'].dt.strftime('%m-%Y')
grupos = df_macro_dados[['Município', 'UVR Número', 'MesAno']].drop_duplicates()
output_dir = os.path.join(SCRIPT_DIR, 'outputs')
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

#GERAÇÃO DAS PLANILHAS
for index, row in grupos.iterrows():
    municipio = row['Município']
    uvr = int(row['UVR Número'])
    mes_ano = row['MesAno']
    
    # Definição do nome e do caminho do arquivo de saída
    output_filename = f"{municipio.replace(' ', '_')}_UVR-{uvr}_{mes_ano}.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    
    print(f"Gerando planilha para: {output_filename}")

    wb = load_workbook(template_path)
    macro_grupo = df_macro_dados[(df_macro_dados['Município'] == municipio) & (df_macro_dados['UVR Número'] == uvr) & (df_macro_dados['MesAno'] == mes_ano)]
    micro_grupo = df_micro_dados[(df_micro_dados['Município'] == municipio) & (df_micro_dados['UVR Número'] == uvr) & (df_micro_dados['MesAno'] == mes_ano)]

    #Macro Dados
    if not macro_grupo.empty:
        ws_macro = wb['Macro Dados']
        despesas_no_grupo = micro_grupo[micro_grupo['Tipo do registro'].astype(str).str.strip().str.lower() == 'despesa']
        soma_operacao, soma_manutencao, soma_fundo_caixa = 0.0, 0.0, 0.0
        for _, despesa_row in despesas_no_grupo.iterrows():
            subtipo = despesa_row['Subtipo do registro']
            valor = despesa_row['Valor Absoluto (R$)']
            if isinstance(subtipo, str) and pd.notna(valor):
                subtipo_lower = subtipo.lower()
                if 'manutenção' in subtipo_lower or 'conserto' in subtipo_lower or 'reforma' in subtipo_lower:
                    soma_manutencao += valor
                elif 'fundo de caixa' in subtipo_lower:
                    soma_fundo_caixa += valor
                else:
                    soma_operacao += valor
        
        ws_macro['A2'] = macro_grupo['Município'].iloc[0]
        ws_macro['B2'] = macro_grupo['UVR Número'].iloc[0]
        ws_macro['C2'] = macro_grupo['Data de referência'].iloc[0].strftime('%m/%Y')
        ws_macro['D2'] = macro_grupo['Número de catadores na UVR no mês da data de referência'].iloc[0]
        ws_macro['E2'] = macro_grupo['Renda média catadores'].iloc[0]
        ws_macro['F2'] = macro_grupo['Observações'].iloc[0]
        ws_macro['G2'] = macro_grupo['Receita venda recicláveis'].iloc[0]
        ws_macro['H2'] = soma_operacao
        ws_macro['I2'] = soma_manutencao
        ws_macro['J2'] = soma_fundo_caixa
        ws_macro['K2'] = macro_grupo['Rejeito (kg)'].iloc[0]

    #Receitas
    ws_receitas = wb['Receitas']
    receitas_por_subtipo = micro_grupo[micro_grupo['Tipo do registro'] == 'Receita da prestação de serviços'].groupby('Subtipo do registro')['Valor Absoluto (R$)'].sum().to_dict()
    
    opcoes_receitas_template = {clean_text(col): col for col in template_receitas_df.columns}
    receitas_mapeadas = {col: 0.0 for col in opcoes_receitas_template.values()}
    
    if DEBUG_MODE: print("--- DEBUG: RECEITAS ---")
    for subtipo_original, valor in receitas_por_subtipo.items():
        if pd.isna(valor): continue
        subtipo_limpo = clean_text(subtipo_original)

        if subtipo_limpo in SYNONYM_MAP:
            match_final = SYNONYM_MAP[subtipo_limpo]
            score = 100
            if DEBUG_MODE: print(f"  - Item: '{subtipo_original}' -> Mapeado via Sinônimo para '{match_final}'")
        else:
            match_limpo, score = process.extractOne(subtipo_limpo, opcoes_receitas_template.keys(), scorer=fuzz.token_set_ratio)
            match_final = opcoes_receitas_template.get(match_limpo)
            if DEBUG_MODE:
                print(f"  - Item: '{subtipo_original}' -> Limpo: '{subtipo_limpo}'")
                print(f"    Melhor Match: '{match_final}' | Score: {score}%")
        
        if score > 80 and match_final:
            receitas_mapeadas[match_final] += valor
        elif 'Outro Tipo' in receitas_mapeadas:
            receitas_mapeadas['Outro Tipo'] += valor
            if DEBUG_MODE: print(f"    -> Ação: Enviado para 'Outro Tipo'.")
            
    for i, col_header in enumerate(template_receitas_df.columns, 1):
        ws_receitas.cell(row=2, column=i).value = receitas_mapeadas.get(col_header, 0)

    #Despesas
    ws_despesas = wb['Despesas']
    opcoes_despesas_template = {clean_text(nome): nome for nome in template_despesas_df['Nome'].dropna()}
    valores_despesas = {nome: 0.0 for nome in opcoes_despesas_template.values()}
    valores_despesas.setdefault('Outras Despesas de Operação', 0.0)
    
    despesas_dados = micro_grupo[micro_grupo['Tipo do registro'].astype(str).str.strip().str.lower() == 'despesa']
    if DEBUG_MODE: print("--- DEBUG: DESPESAS ---")
    for _, row_d in despesas_dados.iterrows():
        desc_original = row_d['Subtipo do registro']
        desc_limpa = clean_text(desc_original)
        valor = row_d['Valor Absoluto (R$)']
        if not desc_limpa or pd.isna(valor): continue

        if desc_limpa in SYNONYM_MAP:
            match_final = SYNONYM_MAP[desc_limpa]
            score = 100
            if DEBUG_MODE: print(f"  - Item: '{desc_original}' -> Mapeado via Sinônimo para '{match_final}'")
        else:
            match_limpo, score = process.extractOne(desc_limpa, opcoes_despesas_template.keys(), scorer=fuzz.token_set_ratio)
            match_final = opcoes_despesas_template.get(match_limpo)
            if DEBUG_MODE:
                print(f"  - Item: '{desc_original}' -> Limpo: '{desc_limpa}'")
                print(f"    Melhor Match: '{match_final}' | Score: {score}%")
        
        if score >= MATCH_THRESHOLD and match_final:
            valores_despesas[match_final] += valor
        else:
            valores_despesas['Outras Despesas de Operação'] += valor
            if DEBUG_MODE: print(f"    -> Ação: Enviado para 'Outras Despesas'.")

    col_valor_idx = list(template_despesas_df.columns).index('Valor') + 1
    for r_idx, nome in enumerate(template_despesas_df['Nome'], 3):
        if pd.notna(nome) and nome in valores_despesas:
            ws_despesas.cell(row=r_idx, column=col_valor_idx).value = valores_despesas[nome]

    # Materiais
    ws_materiais = wb['Materiais']
    opcoes_materiais_template = {clean_text(subtipo): subtipo for subtipo in template_materiais_df['Subtipo'].dropna()}
    mapa_categoria_material = pd.Series(template_materiais_df.Categoria.values, index=template_materiais_df.Subtipo).to_dict()
    valores_materiais = {subtipo: {'Quantidade': 0.0, 'Valor': 0.0} for subtipo in opcoes_materiais_template.values()}
    
    materiais_dados = micro_grupo[micro_grupo['Tipo do registro'] == 'Receita da venda de recicláveis']
    if DEBUG_MODE: print("--- DEBUG: MATERIAIS ---")
    for _, row_m in materiais_dados.iterrows():
        nome_original = row_m['Material']
        nome_limpo = clean_text(nome_original)
        quantidade = row_m.get('Quantidade', 0) if pd.notna(row_m.get('Quantidade')) else 0
        valor = row_m.get('Valor Absoluto (R$)', 0) if pd.notna(row_m.get('Valor Absoluto (R$)')) else 0
        if not nome_limpo: continue

        match_limpo, score = process.extractOne(nome_limpo, opcoes_materiais_template.keys(), scorer=fuzz.token_set_ratio)
        match_final = opcoes_materiais_template.get(match_limpo)
        if DEBUG_MODE:
            print(f"  - Item: '{nome_original}' -> Limpo: '{nome_limpo}'")
            print(f"    Melhor Match: '{match_final}' | Score: {score}%")
        
        if score >= MATCH_THRESHOLD and match_final:
            valores_materiais[match_final]['Quantidade'] += quantidade
            valores_materiais[match_final]['Valor'] += valor
        else:
            categoria = mapa_categoria_material.get(match_final)
            outro_key = f'Outro {categoria}' if categoria else None
            if outro_key not in valores_materiais:
                outro_key = next((k for k in valores_materiais if 'outro' in k.lower()), None)
            if outro_key:
                valores_materiais[outro_key]['Quantidade'] += quantidade
                valores_materiais[outro_key]['Valor'] += valor
                if DEBUG_MODE: print(f"    -> Ação: Enviado para '{outro_key}'.")
    
    col_qtd_idx = list(template_materiais_df.columns).index('Quantidade') + 1
    col_val_idx = list(template_materiais_df.columns).index('Valor') + 1
    for r_idx, subtipo in enumerate(template_materiais_df['Subtipo'], 3):
        if pd.notna(subtipo) and subtipo in valores_materiais:
            ws_materiais.cell(row=r_idx, column=col_qtd_idx).value = valores_materiais[subtipo]['Quantidade']
            ws_materiais.cell(row=r_idx, column=col_val_idx).value = valores_materiais[subtipo]['Valor']

 
    wb.save(output_path)

print("\nProcesso finalizado!")