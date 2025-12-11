# Importações necessárias
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from pathlib import Path
import pandas as pd
from datetime import timedelta
from utils import (
    cabeçalho_fill, cabeçalho_font, enviado_fill, analise_fill, enviado_font,
    semtecnico_fill, atrasado_fill, validado_nao_fill, validado_sim_fill, duplicado_fill, outras_fill, atrasado2_fill,
    cores_regionais, bordas, alinhamento,
    normalizar_texto, normalizar_uvr, aplicar_estilo_status
)


VARIAVEIS_ANALISE = {
    # Mantendo a variável original
    "Receita Vendas (R$)":         {"col_form_idx": 6, "col_media_s1": "E", "col_media_s2": "M"},
    # Adicionando as novas variáveis
    "Receita Serviços (R$)":       {"col_form": "receita_servicos", "col_media_s1": "F", "col_media_s2": "N"},
    "Despesas (R$)":               {"col_form": "despesas", "col_media_s1": "G", "col_media_s2": "O"},
    "Material Reciclado (T)": {"col_form": "material_reciclado", "col_media_s1": "H", "col_media_s2": "P"},
    "Rejeito (T)":                {"col_form": "rejeito", "col_media_s1": "I", "col_media_s2": "Q"},
    "Total Material Processado (T)": {"col_form": "total_material_processado", "col_media_s1": "J", "col_media_s2": "R"},
    "Postos de Trabalho (U)":     {"col_form": "postos_de_trabalho", "col_media_s1": "K", "col_media_s2": "S"},
    "Renda Média (R$)":            {"col_form": "renda_media", "col_media_s1": "L", "col_media_s2": "T"},
}


def col_to_idx(letra):
    return ord(letra.upper()) - ord('A')

# Define os caminhos dos arquivos envolvidos
caminho_script = Path(__file__).resolve()
pasta_scripts = caminho_script.parent
pasta_inputs = pasta_scripts.parent / "inputs"


# Caminho do arquivo do banco e arquivos auxiliares (originais do drive)
csv_file_input = pasta_inputs/"form4.csv"
media_file_input = pasta_inputs/"form4-médias.csv"
planilhas_auxiliares = {
    "belem": pasta_inputs / "0 - Belém" / "0 - Monitoramento Form 4.xlsx",
    "expansao": pasta_inputs / "0 - Expansão" / "0 - Monitoramento Form 4.xlsx",
    "grs": pasta_inputs / "0 - GRS II" / "0 - Monitoramento Form 4.xlsx"
}

# Carrega a planilha principal
df_input = pd.read_csv(csv_file_input, dtype=str)
df_medias = pd.read_csv(media_file_input, dtype=str)

# Dicionários para armazenar os dados extraídos
dados_atualizados = {}
div_por_municipio = {}
regionais_por_municipio = {}
dados_medias = {}

# Processa e armazena as médias
col_municipio_media = df_medias.columns[0]
col_uvr_media = df_medias.columns[2]
col_media_s1 = df_medias.columns[4]
col_media_s2 = df_medias.columns[12]

# carrega todas as médias
for _, row in df_medias.iterrows():
    municipio = row[col_municipio_media]
    uvr_nro = row[col_uvr_media]
    if pd.notna(municipio) and pd.notna(uvr_nro):
        chave = (normalizar_texto(str(municipio)), normalizar_uvr(str(uvr_nro)))
        dados_medias[chave] = {}  # Prepara para receber múltiplas variáveis

        for nome_variavel, config in VARIAVEIS_ANALISE.items():
            idx_s1 = col_to_idx(config["col_media_s1"])
            idx_s2 = col_to_idx(config["col_media_s2"])
            if nome_variavel not in dados_medias[chave]:
                dados_medias[chave][nome_variavel] = {}
            
            dados_medias[chave][nome_variavel] = {
                "media_s1": pd.to_numeric(row.iloc[idx_s1], errors='coerce'),
                "media_s2": pd.to_numeric(row.iloc[idx_s2], errors='coerce')
            }

# Converte a data de referência para o formato "MM.AA"
def converter_data_para_mes_ano(data_referencia):
    if isinstance(data_referencia, datetime):
        return data_referencia.strftime("%m.%y")
    else:
        try:
            return datetime.strptime(data_referencia, "%Y-%m-%d").strftime("%m.%y")
        except (ValueError, TypeError):
            return ""

# Remove caracteres inválidos do nome da aba
def limpar_nome_aba(nome):
    return nome.replace("/", "-").replace("\\", "-").replace(":", "-").replace("*", "-").replace("?", "-").replace("[", "").replace("]", "")


# carrega todos os valores enviados
for _, row in df_input.iterrows():
    municipio = row['gm_nome']
    uvr_nro = row['guvr_numero']
    data_envio = row['data_de_envio']
    tc_uvr = row['nome_tc_uvr']  
    data_referencia = row['data_de_referencia']

    # Pega os valores de todas as variáveis que queremos analisar
    valores_enviados = {}
    for nome_variavel, config in VARIAVEIS_ANALISE.items():
        # Lida com o caso original (por índice) e os novos (por nome de coluna)
        if "col_form_idx" in config:
            coluna = df_input.columns[config["col_form_idx"]]
        else:
            coluna = config["col_form"]
        
        if coluna in row:
            valores_enviados[nome_variavel] = pd.to_numeric(row[coluna], errors='coerce')
        else:
            valores_enviados[nome_variavel] = None


    if isinstance(municipio, str):
        municipio_uvr_normalizado = f"{normalizar_texto(municipio)}_{uvr_nro}"
    else:
        continue

    mes_ano = converter_data_para_mes_ano(data_referencia)

    if isinstance(data_envio, datetime):
        data_envio_formatada = data_envio.strftime("%d/%m/%Y")
    else:
        try:
            data_envio_formatada = datetime.strptime(data_envio, "%Y-%m-%d").strftime("%d/%m/%Y")
        except (ValueError, TypeError):
            data_envio_formatada = ""

    chave = (municipio_uvr_normalizado, mes_ano)
    if chave in dados_atualizados:
        dados_atualizados[chave]["datas_envio"].append(data_envio_formatada)
        dados_atualizados[chave]["status"] = "Duplicado"
    else:
        dados_atualizados[chave] = {
            "datas_envio": [data_envio_formatada],
            "status": "Enviado",
            "municipio_original": municipio,
            "uvr_nro": uvr_nro,
            "mes_ano": mes_ano,
            "tc_uvr" : tc_uvr,
            "valores_enviados": valores_enviados, # <<< AQUI ESTÁ A MUDANÇA PRINCIPAL
            "data_referencia_dt": pd.to_datetime(data_referencia, errors='coerce')
        }

# Cria um novo workbook para cada planilha auxiliar
wb_final = {nome: Workbook() for nome in planilhas_auxiliares}
for nome in wb_final:
    wb_final[nome].remove(wb_final[nome].active)

# Processa cada planilha auxiliar
for nome, caminho in planilhas_auxiliares.items():
    wb_aux = load_workbook(caminho)

    abas_para_copiar = ["Resumo", "Monitoramento", "Regionais"]

    for nome_aba in abas_para_copiar:
        if nome_aba in wb_aux.sheetnames: #verifica se existe, faz isso em todos (grs,expansao,belem)
            print(f"Copiando aba '{nome_aba}' para o arquivo de '{nome}'...")
            ws_origem = wb_aux[nome_aba]
            ws_destino = wb_final[nome].create_sheet(nome_aba)

            # Copia os dados e estilos célula por célula
            for row in ws_origem.iter_rows():
                for cell in row:
                    new_cell = ws_destino.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color)
                        new_cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom)
                        new_cell.fill = PatternFill(fill_type=cell.fill.fill_type, start_color=cell.fill.start_color, end_color=cell.fill.end_color)
                        new_cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=cell.alignment.wrap_text)
                        new_cell.number_format = cell.number_format

            # Copia as dimensões das colunas e linhas
            for col_letter, dim in ws_origem.column_dimensions.items():
                ws_destino.column_dimensions[col_letter].width = dim.width
            for row_index, dim in ws_origem.row_dimensions.items():
                ws_destino.row_dimensions[row_index].height = dim.height

            # Copia as células mescladas
            for merged_cell_range in ws_origem.merged_cells.ranges:
                ws_destino.merge_cells(str(merged_cell_range))
            
            for dv in ws_origem.data_validations.dataValidation:
                ws_destino.add_data_validation(dv)

            for range_string in ws_origem.conditional_formatting:
                rules_list = ws_origem.conditional_formatting[range_string]
                    
                for rule in rules_list:
                    ws_destino.conditional_formatting.add(range_string, rule)

    for aba in wb_aux.sheetnames:
        # Só processa abas no formato MM.AA
        if aba.count('.') == 1 and all(x.isdigit() for x in aba.split('.')):
            mes_ano_aux = aba
            ws_aux = wb_aux[aba]

            mes_ano_limpo = limpar_nome_aba(mes_ano_aux)
            if mes_ano_limpo not in wb_final[nome].sheetnames:
                wb_final[nome].create_sheet(title=mes_ano_limpo)

            ws_final = wb_final[nome][mes_ano_limpo]

            dv_sim_nao = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
            ws_final.add_data_validation(dv_sim_nao)

            dv_sim_nao_ti = DataValidation(type="list", formula1='"Sim,Não,Em Análise"', allow_blank=True)
            ws_final.add_data_validation(dv_sim_nao_ti)

            dv_status = DataValidation(type="list", formula1='"Enviado, Atrasado, Atrasado >= 2, Outras Ocorrências, Sem Técnico, Duplicado"', allow_blank=True) #dropdown com sim e nao
            ws_final.add_data_validation(dv_status)           

            # Copia cabeçalhos com formatação
            headers = [cell.value for cell in ws_aux[1]]
            for col_num, header in enumerate(headers, start=1):
                cell = ws_final.cell(row=1, column=col_num, value=header)
                cell.fill = cabeçalho_fill
                cell.font = cabeçalho_font
                cell.border = bordas
                cell.alignment = alinhamento

            ws_final.auto_filter.ref = f"A1:G1"

            # Calcula mês/ano atual
            hoje = datetime.today()
            mes_atual = hoje.month
            ano_atual = hoje.year

            # Função auxiliar para calcular a diferença de meses
            def diferenca_em_meses(ano_alvo, mes_alvo, ano_base, mes_base):
                return (ano_base - ano_alvo) * 12 + (mes_base - mes_alvo)

            # Processa linhas de dados
            for row_idx, row in enumerate(ws_aux.iter_rows(min_row=2, values_only=True), start=2):
                regional = row[0]
                municipio_original = row[1]
                uvr_nro_original = row[2]
                row_data = list(row)


                if aba != '01.25':
                    formula = (
                        f'=IFERROR(IF(INDEX(\'01.25\'!D2:D500, '
                        f'MATCH(B{row_idx}&C{row_idx}, INDEX(\'01.25\'!B2:B500&\'01.25\'!C2:C500, 0), 0))="", "", '
                        f'INDEX(\'01.25\'!D2:D500, '
                        f'MATCH(B{row_idx}&C{row_idx}, INDEX(\'01.25\'!B2:B500&\'01.25\'!C2:C500, 0), 0))), "")'
                    )

                    
                    row_data[3] = formula
               
                if not isinstance(municipio_original, str) or not municipio_original.strip():
                    continue

                municipio_uvr_normalizado = f"{normalizar_texto(municipio_original)}_{normalizar_uvr(uvr_nro_original)}"
                chave_busca = (municipio_uvr_normalizado, mes_ano_aux)
                div_por_municipio[municipio_uvr_normalizado] = nome
                regionais_por_municipio[municipio_uvr_normalizado] = regional

                situacao_atual = row_data[4]
                tem_envio_existente = bool(row_data[5]) and isinstance(row_data[5], str) and row_data[5].strip()

                # Atualiza dados conforme a planilha principal
                if chave_busca in dados_atualizados:

                    # 1. Contar o número de datas de envio na planilha anterior (coluna F, índice 5)
                    datas_antigas_str = row[5] if row[5] and isinstance(row[5], str) else ""
                    num_datas_antigas = len([data for data in datas_antigas_str.split(',') if data.strip()])

                    # 2. Contar o número de novas datas de envio a partir dos dados do CSV
                    num_datas_novas = len(dados_atualizados[chave_busca]["datas_envio"])

                    # 3. Se o número de datas aumentou, marcar "Validado pelo Regional" (coluna G, índice 6) como "Não"
                    if num_datas_novas > num_datas_antigas:
                        row_data[6] = "Não"
                    
                    row_data[5] = ", ".join(dados_atualizados[chave_busca]["datas_envio"])
                    row_data[4] = dados_atualizados[chave_busca]["status"]
                elif not tem_envio_existente:
                    if situacao_atual not in ("Sem Técnico", "Outras Ocorrências"):
                        try:
                            aba_mes, aba_ano = map(int, mes_ano_aux.split("."))
                            aba_ano += 2000
                        except:
                            aba_mes, aba_ano = None, None

                        if aba_ano and aba_mes:
                            diff = diferenca_em_meses(aba_ano, aba_mes, ano_atual, mes_atual)
                            if diff == 1:
                                row_data[4] = "Atrasado"
                            elif diff >= 2:
                                row_data[4] = "Atrasado >= 2"

                # Estiliza as linhas
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_final.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = bordas
                    cell.alignment = alinhamento
                    cell.font = Font(name='Arial', size=11)
                    
                    if col_idx == 7:
                        dv_sim_nao.add(cell.coordinate)

                    if col_idx == 10:
                        dv_sim_nao_ti.add(cell.coordinate)

                    if col_idx == 5:
                        dv_status.add(cell.coordinate) 
                    

                # Aplica cor para célula de validação
                if row_data[6] == "Não":
                    ws_final.cell(row=row_idx, column=7).fill = validado_nao_fill
                elif row_data[6] == "Sim":
                    ws_final.cell(row=row_idx, column=7).fill = validado_sim_fill

                # Aplica cor da regional
                if regional in cores_regionais:
                    cor_hex = cores_regionais[regional]
                    ws_final.cell(row=row_idx, column=1).fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")

            # Aplica estilo com base no status
            for row_idx in range(2, ws_final.max_row + 1):
                status_cell = ws_final.cell(row=row_idx, column=5)
                status = status_cell.value
                aplicar_estilo_status(status_cell, status)

            # Ajusta largura das colunas
            for col in ws_final.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                col_letter = col[0].column_letter
                wb_final[nome][mes_ano_limpo].column_dimensions[col_letter].width = max_length + 5

            coluna_validado_regional = f"G2:G{ws_final.max_row}" # Coluna G é a 7ª coluna
            coluna_validado_ti = f"J2:J{ws_final.max_row}" # Coluna G é a 10ª coluna
            coluna_status = f"E2:E{ws_final.max_row}" # Coluna E é a 5ª coluna

            rule_sim = CellIsRule(operator='equal', formula=['"Sim"'], stopIfTrue=True, fill=validado_sim_fill)
            ws_final.conditional_formatting.add(coluna_validado_regional, rule_sim) #Se for selecionado Sim, pinta de verde
            ws_final.conditional_formatting.add(coluna_validado_ti, rule_sim) #Se for selecionado Sim, pinta de verde

            rule_nao = CellIsRule(operator='equal', formula=['"Não"'], stopIfTrue=True, fill=validado_nao_fill)
            ws_final.conditional_formatting.add(coluna_validado_regional, rule_nao) #Se for selecionado Não, pinta de vermelho
            ws_final.conditional_formatting.add(coluna_validado_ti, rule_nao) #Se for selecionado Sim, pinta de verde

            rule_analise = CellIsRule(operator='equal', formula=['"Em Análise"'], stopIfTrue=True, fill=analise_fill)  
            ws_final.conditional_formatting.add(coluna_validado_ti, rule_analise)    

            status_rules = {
            "Enviado": {"fill": enviado_fill, "font": enviado_font},
            "Atrasado": {"fill": atrasado_fill, "font": enviado_font},
            "Atrasado >= 2": {"fill": atrasado2_fill, "font": enviado_font},
            "Outras Ocorrências": {"fill": outras_fill, "font": enviado_font},
            "Sem Técnico": {"fill": semtecnico_fill, "font": enviado_font},
            "Duplicado": {"fill": duplicado_fill, "font": enviado_font}
        }

            for status_text, styles in status_rules.items():
                rule = CellIsRule(operator='equal',
                                formula=[f'"{status_text}"'],
                                stopIfTrue=True,
                                fill=styles["fill"],
                                font=styles["font"])
                ws_final.conditional_formatting.add(coluna_status, rule) 

            ws_final.freeze_panes = 'D1' #Congela as colunas A,B,C  
            ws_final.column_dimensions['D'].width = 45



 # processa a aba de irregulares (grs,expansao e belem)
for nome, wb in wb_final.items():

    print(nome)
    chaves_existentes = set()
    
    caminho_aux = planilhas_auxiliares[nome]
    wb_aux = load_workbook(caminho_aux)

    # cria a aba de irregulares no arquivo final (ela sempre é recriada, porém coletando as informações já existentes na planilha de entrada)
    if "Irregulares" in wb.sheetnames:
        wb.remove(wb["Irregulares"]) # remove qualquer possível versão antiga para evitar conflitos
    aba_irregulares_final = wb.create_sheet("Irregulares")


    colunas_irregulares_padrao = [
        "Regional", "Município", "UVR", "Técnico de UVR", 
        "Data de Envio", "Mês de referência", "Validado pelo Regional", "Observações", "Formulários para Deletar (ID)", "Validado Equipe de TI", "Resposta Equipe de TI"
    ]

    
    # Escreve o novo cabeçalho 
    for col_num, col_name in enumerate(colunas_irregulares_padrao, start=1):
        cell = aba_irregulares_final.cell(row=1, column=col_num, value=col_name)
        cell.fill = cabeçalho_fill
        cell.font = cabeçalho_font
        cell.border = bordas
        cell.alignment = alinhamento

    # primeira etapa: migrar dados da aba de irregulares do arquivo de entrada
    if "Irregulares" in wb_aux.sheetnames:
        aba_irregulares_origem = wb_aux["Irregulares"]

        # cria um conjunto com as chaves de todos os novos envios para verificação
        chaves_novos_envios = set()
        for chave_composta, info in dados_atualizados.items():
            _municipio_uvr, mes_ano = chave_composta
            for data_envio in info["datas_envio"]:
                chave = (
                    normalizar_texto(info["municipio_original"]),
                    normalizar_uvr(info["uvr_nro"]),
                    data_envio,
                    mes_ano
                )
                chaves_novos_envios.add(chave)
        
        headers_origem = [cell.value for cell in aba_irregulares_origem[1]] # captura os nomes dos cabeçalhos da primeira linha da aba de origem
        try:
            # mapeia o índice de cada coluna esperada, conforme a lista de colunas padrão
            idx_map = {h: headers_origem.index(h) for h in colunas_irregulares_padrao if h in headers_origem} 
        except ValueError as e:
            print(f"AVISO: A aba 'Irregulares' em '{caminho_aux}' não tem a coluna esperada")
            idx_map = {}

        if idx_map:
            for row_origem in aba_irregulares_origem.iter_rows(min_row=2, values_only=True):
                municipio = row_origem[idx_map.get("Município")]
                if not municipio: continue

                # Cria uma chave para a linha atual do arquivo de entrada para comparação
                chave_origem = (
                    normalizar_texto(municipio), 
                    normalizar_uvr(row_origem[idx_map.get("UVR")]), 
                    row_origem[idx_map.get("Data de Envio")], 
                    row_origem[idx_map.get("Mês de referência")]
                )
                
                # migra a linha somente se a chave de origem existir nos novos envios
                if chave_origem in chaves_novos_envios:
                    idx_validado_regional = idx_map.get("Validado pelo Regional")
                    valor_validado = row_origem[idx_validado_regional] if idx_validado_regional is not None else "Não"
                    validado = "Sim" if valor_validado == "Sim" else "Não"

                    idx_validado_ti = idx_map.get("Validado Equipe de TI")
                    valor_validado_ti = row_origem[idx_validado_ti] if idx_validado_ti is not None else "Não"
                    validado_TI = "Sim" if valor_validado_ti == "Sim" else "Não"
                    
                    linha_migrada = [
                        row_origem[idx_map.get("Regional", "")] if "Regional" in idx_map else "",
                        municipio,
                        row_origem[idx_map.get("UVR", "")] if "UVR" in idx_map else "",
                        row_origem[idx_map.get("Técnico de UVR", "")] if "Técnico de UVR" in idx_map else "",
                        row_origem[idx_map.get("Data de Envio", "")] if "Data de Envio" in idx_map else "",
                        row_origem[idx_map.get("Mês de referência", "")] if "Mês de referência" in idx_map else "",
                        validado,
                        row_origem[idx_map.get("Observações", "")] if "Observações" in idx_map else "",
                        row_origem[idx_map.get("Formulários para Deletar (ID)", "")] if "Formulários para Deletar (ID)" in idx_map else "",
                        validado_TI, 
                        row_origem[idx_map.get("Resposta Equipe de TI", "")] if "Resposta Equipe de TI" in idx_map else ""                   
                    ]
                    aba_irregulares_final.append(linha_migrada)

                    # Adiciona a chave da linha migrada para evitar duplicatas na segunda etapa
                    chaves_existentes.add(chave_origem)

    # segunda etapa: adicionar novos registros irregulares do csv que ainda não existem
    for chave_composta, info in dados_atualizados.items():
        municipio_uvr, mes_ano = chave_composta
        
        if mes_ano not in wb.sheetnames and div_por_municipio.get(municipio_uvr) == nome: # verifica se é irregular
            for data_envio in info["datas_envio"]:
                chave_nova = (
                    normalizar_texto(info["municipio_original"]),
                    normalizar_uvr(info["uvr_nro"]),
                    data_envio,
                    mes_ano
                )
                
                if chave_nova not in chaves_existentes: #verifica se a chave já não existe
                    nova_linha_dados = [
                        regionais_por_municipio.get(municipio_uvr, ""),
                        info["municipio_original"], 
                        info["uvr_nro"], 
                        info["tc_uvr"],
                        data_envio, 
                        mes_ano, 
                        "Não", 
                        "", 
                        "",
                        "Não",
                        "",
                    ]
                    aba_irregulares_final.append(nova_linha_dados)
                    chaves_existentes.add(chave_nova)

    # aplicar estilização na aba de irregulares
    for row_idx in range(2, aba_irregulares_final.max_row + 1):
        for col_idx in range(1, len(colunas_irregulares_padrao) + 1):
            cell = aba_irregulares_final.cell(row=row_idx, column=col_idx)
            cell.border = bordas
            cell.alignment = alinhamento
            cell.font = Font(name='Arial', size=11)
        
        regional_cell = aba_irregulares_final.cell(row=row_idx, column=1)
        if regional_cell.value in cores_regionais:
            cor_hex = cores_regionais[regional_cell.value]
            regional_cell.fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")

        status_cell = aba_irregulares_final.cell(row=row_idx, column=5)
        aplicar_estilo_status(status_cell, status_cell.value)
    
    
    if aba_irregulares_final.max_row > 1:
        # Cria o dropdown
        dv_sim_nao_irr = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=False)
        aba_irregulares_final.add_data_validation(dv_sim_nao_irr)
        
        # Define o range da coluna a ser afetada (H2 até a última linha)
        range_validado = f"G2:G{aba_irregulares_final.max_row}"
        range_validado_TI = f"J2:J{aba_irregulares_final.max_row}"        
        dv_sim_nao_irr.add(range_validado)
        dv_sim_nao_irr.add(range_validado_TI)

        # Define as regras de formatação condicional
        rule_sim_irr = CellIsRule(operator='equal', formula=['"Sim"'], stopIfTrue=True, fill=validado_sim_fill)
        rule_nao_irr = CellIsRule(operator='equal', formula=['"Não"'], stopIfTrue=True, fill=validado_nao_fill)


        # Aplica as regras ao range
        aba_irregulares_final.conditional_formatting.add(range_validado, rule_sim_irr)
        aba_irregulares_final.conditional_formatting.add(range_validado, rule_nao_irr)

        aba_irregulares_final.conditional_formatting.add(range_validado_TI, rule_sim_irr)
        aba_irregulares_final.conditional_formatting.add(range_validado_TI, rule_nao_irr)

    # Ajusta a largura das colunas
    if aba_irregulares_final.max_row > 1:
        for col in aba_irregulares_final.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            aba_irregulares_final.column_dimensions[col[0].column_letter].width = max_length + 5

    aba_irregulares_final.freeze_panes = 'D1'
    aba_irregulares_final.auto_filter.ref = f"A1:G1"

# --- Lógica para a Aba "Discrepantes" (Versão Matriz "Wide") ---
for nome, wb in wb_final.items():
    print(f"Processando Discrepantes para '{nome}'...")

    # --- 1. Ler dados existentes da aba "Discrepantes" ---
    dados_antigos_discrepantes = {}
    caminho_aux = planilhas_auxiliares[nome]
    wb_aux = load_workbook(caminho_aux)
    if "Discrepantes" in wb_aux.sheetnames:
        ws_antiga = wb_aux["Discrepantes"]
        headers_antigos = [cell.value for cell in ws_antiga[1]]
        
        try:
            col_municipio_idx = headers_antigos.index("Município")
            col_uvr_idx = headers_antigos.index("UVR")
            col_mes_ref_idx = headers_antigos.index("Mês Referência")
            col_validado_idx = headers_antigos.index("Validado pelo Regional")
            col_obs_idx = headers_antigos.index("Observações")
            col_tecnico_idx = headers_antigos.index("Técnico UVR")       
            col_data_envio_idx = headers_antigos.index("Data de Envio")
        except ValueError:
            print(f"Aviso: A aba 'Discrepantes' em '{caminho_aux}' não possui as colunas esperadas. Os dados de validação não serão migrados.")
            col_municipio_idx = -1

        if col_municipio_idx != -1:
            for row in ws_antiga.iter_rows(min_row=2, values_only=True):
                municipio = row[col_municipio_idx]
                uvr = row[col_uvr_idx]
                mes_ref = row[col_mes_ref_idx]
                
                if municipio and uvr and mes_ref:
                    chave = (normalizar_texto(str(municipio)), normalizar_uvr(str(uvr)), str(mes_ref))
                    
                    dados_antigos_discrepantes[chave] = {
                        "validado": row[col_validado_idx],
                        "observacoes": row[col_obs_idx],
                        "data_envio": row[col_data_envio_idx],
                        "tecnico":row[col_tecnico_idx],
                        "valores_antigos": {var: row[headers_antigos.index(var)] for var in VARIAVEIS_ANALISE.keys() if var in headers_antigos}
                    }


    if "Discrepantes" in wb.sheetnames:
        wb.remove(wb["Discrepantes"])
    ws_discrepantes = wb.create_sheet("Discrepantes")

    # --- 2. Construção do cabeçalho dinâmico ---
    colunas_comuns = ["Regional", "Município", "UVR", "Técnico UVR", "Mês Referência", "Data de Envio"]
    nomes_variaveis = list(VARIAVEIS_ANALISE.keys())
    headers = colunas_comuns + nomes_variaveis + ["Validado pelo Regional", "Observações"]
    
    for col_num, header_text in enumerate(headers, start=1):
        cell = ws_discrepantes.cell(row=1, column=col_num, value=header_text)
        cell.fill = cabeçalho_fill
        cell.font = cabeçalho_font
        cell.border = bordas
        cell.alignment = alinhamento

    # --- 3. lógica de coleta de dados ---
    abas_mensais_existentes = set(wb.sheetnames)
    discrepantes_data = []
    chaves_discrepantes_atuais = set()

    hoje = datetime.today()
    ano_atual = hoje.year
    semestre_atual = 1 if 1 <= hoje.month <= 6 else 2
    
    # Loop 1: Itera sobre cada envio do formulário para encontrar discrepâncias atuais
    for chave_composta, info in dados_atualizados.items():
        municipio_uvr, mes_ano_envio = chave_composta
        if div_por_municipio.get(municipio_uvr) == nome and mes_ano_envio in abas_mensais_existentes:
            data_ref = info.get("data_referencia_dt")
            if not pd.notna(data_ref): continue

            ano_ref = data_ref.year
            semestre_ref = 1 if 1 <= data_ref.month <= 6 else 2
            is_valid_semester = False
            if semestre_atual == 1:
                if (ano_ref == ano_atual and semestre_ref == 1) or (ano_ref == ano_atual - 1 and semestre_ref == 2): is_valid_semester = True
            else: 
                if ano_ref == ano_atual: is_valid_semester = True
            if not is_valid_semester: continue

            tem_alguma_discrepancia = False
            resultados_do_envio = {}

            # Loop 2: Analisa todas as variáveis para o envio atual
            for nome_variavel, config in VARIAVEIS_ANALISE.items():
                chave_media = (normalizar_texto(info["municipio_original"]), normalizar_uvr(info["uvr_nro"]))
                valor_enviado = info["valores_enviados"].get(nome_variavel)
                resultados_do_envio[nome_variavel] = {"valor": "-", "desvio": 0}

                if chave_media in dados_medias and pd.notna(valor_enviado) and valor_enviado != 0:
                    medias_da_variavel = dados_medias[chave_media].get(nome_variavel)
                    if medias_da_variavel:
                        media_ref = 0
                        if semestre_atual == 2:
                            media_ref = medias_da_variavel["media_s1"] if semestre_ref == 1 else medias_da_variavel["media_s2"]
                        else:
                            media_ref = medias_da_variavel["media_s2"] if semestre_ref == 1 else medias_da_variavel["media_s1"]
                        
                        if pd.notna(media_ref) and media_ref != 0:
                            desvio = abs((valor_enviado - media_ref) / media_ref) * 100
                            if desvio >= 60:
                                tem_alguma_discrepancia = True
                                resultados_do_envio[nome_variavel] = {"valor": valor_enviado, "desvio": desvio}

            if tem_alguma_discrepancia:
                chave_atual = (normalizar_texto(info["municipio_original"]), normalizar_uvr(info["uvr_nro"]), mes_ano_envio)
                chaves_discrepantes_atuais.add(chave_atual)

                validado = "Não"
                observacoes = ""

             
                novos_valores_discrepantes = {
                    k: round(v['valor'], 2) for k, v in resultados_do_envio.items() 
                    if v['desvio'] >= 60 and isinstance(v['valor'], (int, float))
                }

                if chave_atual in dados_antigos_discrepantes:
                    dados_antigos = dados_antigos_discrepantes[chave_atual]
                    observacoes = dados_antigos.get("observacoes", "") # Mantém as observações para referência
                    
                    antigos_valores_discrepantes = {
                        k: round(v, 2) for k, v in dados_antigos["valores_antigos"].items()
                        if isinstance(v, (int, float))
                    }
                    
                    if antigos_valores_discrepantes == novos_valores_discrepantes:
                        validado = dados_antigos.get("validado", "Não")
                    # Se forem diferentes, 'validado' permanece "Não", indicando que precisa de re-validação.
                

                linha_de_dados = {
                    "regional": regionais_por_municipio.get(municipio_uvr, ""),
                    "municipio": info["municipio_original"],
                    "uvr": info["uvr_nro"],
                    "tc_uvr": info.get("tc_uvr", ""),
                    "mes_ano": mes_ano_envio,
                    "data_envio": info.get("datas_envio", [""])[0],
                    "resultados": resultados_do_envio,
                    "validado": validado,
                    "observacoes": observacoes
                }
                discrepantes_data.append(linha_de_dados)

    # --- 4. Adicionar itens corrigidos ---
    for chave_antiga, dados_antigos in dados_antigos_discrepantes.items():
        if chave_antiga not in chaves_discrepantes_atuais and dados_antigos.get("validado") != "Corrigido":
            municipio_norm, uvr_norm, mes_ano = chave_antiga
            
            mun_original = ""
            uvr_original = ""
            # Tenta encontrar o nome original do município e uvr
            for k,v in div_por_municipio.items():
                if k == f"{municipio_norm}_{uvr_norm}":
                    # Busca nos dados atualizados para pegar o nome original
                    for info in dados_atualizados.values():
                        if normalizar_texto(info['municipio_original']) == municipio_norm and normalizar_uvr(info['uvr_nro']) == uvr_norm:
                            mun_original = info['municipio_original']
                            uvr_original = info['uvr_nro']
                            break
                    break
            


            discrepantes_data.append({
                "regional": regionais_por_municipio.get(f"{municipio_norm}_{uvr_norm}", ""),
                "municipio": mun_original or municipio_norm, # Fallback para o nome normalizado
                "uvr": uvr_original or uvr_norm, # Fallback
                "tc_uvr": dados_antigos.get("tecnico", ""),
                "mes_ano": mes_ano,
                "data_envio": dados_antigos.get("data_envio", ""),
                "resultados": {nome_variavel: {"valor": dados_antigos["valores_antigos"].get(nome_variavel, "-"), "desvio": 0} for nome_variavel in VARIAVEIS_ANALISE.keys()},
                "validado": "Corrigido",
                "observacoes": dados_antigos.get("observacoes", "")
            })


    # --- 5. lógica de escrita na planilha ---
    discrepantes_data.sort(key=lambda x: (x["municipio"], x["uvr"], x["mes_ano"])) #ordenação

    mapa_colunas_variaveis = {nome_var: i + len(colunas_comuns) + 1 for i, nome_var in enumerate(nomes_variaveis)}
    col_validado_idx = len(headers) - 1
    col_obs_idx = len(headers)

    for i, data in enumerate(discrepantes_data):
        row_idx = i + 2 

        # Passo 1: Obter o nome da regional da linha de dados atual
        nome_da_regional = data["regional"]
        
        # Passo 2: Escreve o valor na célula e guarda a referência da célula
        cell_regional = ws_discrepantes.cell(row=row_idx, column=1, value=nome_da_regional)

        # Passo 3: Busca a cor no dicionário e aplica o preenchimento
        cor_hex = cores_regionais.get(nome_da_regional) 
        if cor_hex:
            fill_regional = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")
            cell_regional.fill = fill_regional
        
        ws_discrepantes.cell(row=row_idx, column=1, value=data["regional"])
        ws_discrepantes.cell(row=row_idx, column=2, value=data["municipio"])
        ws_discrepantes.cell(row=row_idx, column=3, value=data["uvr"])
        ws_discrepantes.cell(row=row_idx, column=4, value=data["tc_uvr"])
        ws_discrepantes.cell(row=row_idx, column=5, value=data["mes_ano"])
        ws_discrepantes.cell(row=row_idx, column=6, value=data["data_envio"])
        ws_discrepantes.cell(row=row_idx, column=col_validado_idx, value=data["validado"])
        ws_discrepantes.cell(row=row_idx, column=col_obs_idx, value=data["observacoes"])

        for col_idx in list(range(1, len(colunas_comuns) + 1)) + [col_validado_idx, col_obs_idx]:
            ws_discrepantes.cell(row=row_idx, column=col_idx).border = bordas
            ws_discrepantes.cell(row=row_idx, column=col_idx).alignment = alinhamento
        
        for nome_variavel, config_col in mapa_colunas_variaveis.items():
            resultado = data["resultados"][nome_variavel]
            valor_a_escrever = resultado["valor"]
            desvio_da_celula = resultado["desvio"]
            
            cell = ws_discrepantes.cell(row=row_idx, column=config_col, value=valor_a_escrever)
            cell.border = bordas
            cell.alignment = alinhamento

            if isinstance(valor_a_escrever, (int, float)):
                cell.number_format = '#,##0.00'

            desvio_fill = None
            if desvio_da_celula >= 80: desvio_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            elif desvio_da_celula >= 70: desvio_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            elif desvio_da_celula >= 60: desvio_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            if desvio_fill:
                cell.fill = desvio_fill

    # --- 6. Adicionar Dropdown e Formatação Condicional ---
    if ws_discrepantes.max_row > 1:
        dv_validacao = DataValidation(type="list", formula1='"Sim,Não,Corrigido"', allow_blank=True)
        ws_discrepantes.add_data_validation(dv_validacao)
        
        range_validado = f"{ws_discrepantes.cell(row=2, column=col_validado_idx).column_letter}2:{ws_discrepantes.cell(row=ws_discrepantes.max_row, column=col_validado_idx).column_letter}{ws_discrepantes.max_row}"
        dv_validacao.add(range_validado)

        rule_sim = CellIsRule(operator='equal', formula=['"Sim"'], stopIfTrue=True, fill=validado_sim_fill)
        rule_nao = CellIsRule(operator='equal', formula=['"Não"'], stopIfTrue=True, fill=validado_nao_fill)
        rule_corrigido = CellIsRule(operator='equal', formula=['"Corrigido"'], stopIfTrue=True, fill=validado_sim_fill)

        ws_discrepantes.conditional_formatting.add(range_validado, rule_sim)
        ws_discrepantes.conditional_formatting.add(range_validado, rule_nao)
        ws_discrepantes.conditional_formatting.add(range_validado, rule_corrigido)


    # Ajusta a largura das colunas
    for col_num in range(1, len(headers) + 1):
        col_letter = ws_discrepantes.cell(row=1, column=col_num).column_letter
        max_length = 0
        for row_num in range(1, ws_discrepantes.max_row + 1):
             cell_value = ws_discrepantes.cell(row=row_num, column=col_num).value
             if cell_value:
                 max_length = max(max_length, len(str(cell_value)))
        adjusted_width = max_length + 10
        ws_discrepantes.column_dimensions[col_letter].width = adjusted_width
        
    ws_discrepantes.freeze_panes = 'D1'
    if ws_discrepantes.max_row > 1:
        last_col_letter = ws_discrepantes.cell(row=1, column=len(headers)).column_letter
        ws_discrepantes.auto_filter.ref = f"A1:{last_col_letter}{ws_discrepantes.max_row}"

    
# ####ANALISE##################

# # --- 4. ANÁLISE ESTATÍSTICA DOS DESVIOS POR INDICADOR ---

# # Dicionário para armazenar a contagem de desvios em faixas distintas
# analise_por_indicador = {
#     nome_var: {
#         'faixa_60_a_70': 0,  # Desvios no intervalo [60%, 70%)
#         'faixa_70_a_80': 0,  # Desvios no intervalo [70%, 80%)
#         'faixa_80_mais': 0   # Desvios no intervalo [80%, infinito)
#     } for nome_var in VARIAVEIS_ANALISE.keys()
# }

# # Loop para popular as contagens
# for data in discrepantes_data:
#     resultados = data.get("resultados", {})
#     for nome_variavel, resultado in resultados.items():
#         desvio = resultado.get("desvio", 0)

#         # Classifica o desvio em uma das três faixas
#         if desvio >= 80:
#             analise_por_indicador[nome_variavel]['faixa_80_mais'] += 1
#         elif desvio >= 70:
#             analise_por_indicador[nome_variavel]['faixa_70_a_80'] += 1
#         elif desvio >= 60:
#             analise_por_indicador[nome_variavel]['faixa_60_a_70'] += 1

# # --- 5. EXIBIÇÃO DOS RESULTADOS NO TERMINAL ---

# # --- 5. EXIBIÇÃO DOS RESULTADOS NO TERMINAL (VERSÃO COM FAIXAS EXCLUSIVAS) ---

# print("\n\n" + "="*70)
# print("--- ANÁLISE PERCENTUAL DE DESVIOS POR FAIXA EXCLUSIVA ---")
# print("="*70)

# # Loop para calcular e imprimir os resultados de cada indicador
# for nome_variavel, contagens in analise_por_indicador.items():
    
#     # Contagens por faixa (estas são as 'fatias' exclusivas)
#     casos_60_a_70 = contagens['faixa_60_a_70']
#     casos_70_a_80 = contagens['faixa_70_a_80']
#     casos_80_mais = contagens['faixa_80_mais']
    
#     # Total de discrepâncias para este indicador
#     total_discrepancias = casos_60_a_70 + casos_70_a_80 + casos_80_mais

#     print(f"\n▶ INDICADOR: {nome_variavel}")
    
#     if total_discrepancias > 0:
#         # Cálculo dos percentuais para cada faixa exclusiva
#         perc_80_mais = (casos_80_mais / total_discrepancias) * 100
#         perc_70_a_80 = (casos_70_a_80 / total_discrepancias) * 100
#         perc_60_a_70 = (casos_60_a_70 / total_discrepancias) * 100

#         print(f"  Total de discrepâncias (desvio >= 60%): {total_discrepancias} casos.")
#         print("-" * 50)
#         print(f"  - Faixa >= 80%:".ljust(35) + f"{casos_80_mais} casos ({perc_80_mais:.2f}%)")
#         print(f"  - Faixa entre 70% e 79.9%:".ljust(35) + f"{casos_70_a_80} casos ({perc_70_a_80:.2f}%)")
#         print(f"  - Faixa entre 60% e 69.9%:".ljust(35) + f"{casos_60_a_70} casos ({perc_60_a_70:.2f}%)")
#     else:
#         print("  - Nenhuma discrepância encontrada.")

# print("\n" + "="*70)

# Salva os novos arquivos com nome atualizado
for nome, wb in wb_final.items():
    novo_caminho = pasta_scripts.parent / "form4" / f"{nome}_atualizado_form4v2.xlsx"
    wb.save(novo_caminho)
    print(f"{novo_caminho} gerado com sucesso")
