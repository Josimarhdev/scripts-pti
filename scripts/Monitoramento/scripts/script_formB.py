import pandas as pd
from openpyxl import Workbook
# NOVO: Importações para Formatação Condicional e Estilos Diferenciais
from openpyxl.formatting.rule import Rule
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
import unicodedata


try:
    from utils import (
        cabeçalho_fill, cabeçalho_font, enviado_fill, enviado_font,
        atrasado_fill, cores_regionais, bordas, alinhamento
    )

    try:
        from utils import nao_possui_fill, nao_possui_font
    except ImportError:
        print("AVISO: Estilos 'nao_possui_fill' e 'nao_possui_font' não encontrados em utils.py.")
        print("Usando estilos padrão (cinza).")
        nao_possui_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        nao_possui_font = Font(name='Calibri', size=11, bold=False, color="000000")

except ImportError:
    print("ERRO: O arquivo 'utils.py' não foi encontrado na pasta 'scripts'.")
    print("Por favor, certifique-se de que o seu arquivo utils.py está no lugar certo.")
    exit()


caminho_script = Path(__file__).resolve()
pasta_scripts = caminho_script.parent
pasta_inputs = pasta_scripts.parent / "inputs"
pasta_outputs = pasta_scripts.parent / "outputs"
pasta_outputs.mkdir(exist_ok=True)


caminhos_csv_formB = {
    "belem": pasta_inputs / "formB-belém.csv",
    "expansao": pasta_inputs / "formB-expansão.csv",
    "grs": pasta_inputs / "formB-grs.csv"
}


caminhos_regionais_xlsx = {
    "belem": pasta_inputs / "0 - Belém" / "0 - Monitoramento Form 4.xlsx",
    "expansao": pasta_inputs / "0 - Expansão" / "0 - Monitoramento Form 4.xlsx",
    "grs": pasta_inputs / "0 - GRS II" / "0 - Monitoramento Form 4.xlsx"
}

# --- FUNÇÕES AUXILIARES --- 
def normalizar_texto(texto: str) -> str:
    if not isinstance(texto, str):
        return texto
    texto_sem_acento = ''.join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    )
    return texto_sem_acento.lower().strip()

def formatar_data(data):
    if str(data).strip() in ['---', '']:
        return '---'
    try:
        return pd.to_datetime(data, dayfirst=True).strftime('%d/%m/%Y')
    except (ValueError, TypeError):
        return data

# --- FUNÇÃO DE GERAÇÃO DA PLANILHA --- 
def gerar_planilha_estilizada_formB(df, nome_convenio):
    print(f"  - Criando Workbook para '{nome_convenio}' (Form B)...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Monitoramento Form B"

    colunas_finais = [
        'Regional', 'Município', 'UVR',
        'Responsável PS', 'Situação PS', 'Data Envio PS',
        'Responsável LR', 'Situação LR', 'Data Envio LR',
        'Responsável OS', 'Situação OS', 'Data Envio OS'
    ]
    
    for col_num, header in enumerate(colunas_finais, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = cabeçalho_fill
        cell.font = cabeçalho_font
        cell.border = bordas
        cell.alignment = alinhamento

    dv = DataValidation(type="list", formula1='"Enviado,Não Enviado,Não Possui"', allow_blank=True)
    ws.add_data_validation(dv)

    for index, row_data in df.iterrows():
        row_num = index + 2
        for col_num, col_name in enumerate(colunas_finais, 1):
            cell = ws.cell(row=row_num, column=col_num, value=row_data[col_name])
            cell.border = bordas
            cell.alignment = alinhamento
            
            if col_name in ['Situação PS', 'Situação LR', 'Situação OS']:
                dv.add(cell)

    dxf_enviado = DifferentialStyle(font=enviado_font, fill=enviado_fill)
    dxf_atrasado = DifferentialStyle(font=enviado_font, fill=atrasado_fill)
    dxf_nao_possui = DifferentialStyle(font=nao_possui_font, fill=nao_possui_fill)
    
    rule_enviado = Rule(type="cellIs", operator="equal", formula=['"Enviado"'], dxf=dxf_enviado)
    rule_atrasado = Rule(type="cellIs", operator="equal", formula=['"Não Enviado"'], dxf=dxf_atrasado)
    rule_nao_possui = Rule(type="cellIs", operator="equal", formula=['"Não Possui"'], dxf=dxf_nao_possui)
    
    max_row = ws.max_row
    colunas_situacao = ['E', 'H', 'K']
    
    for col_letra in colunas_situacao:
        range_str = f"{col_letra}2:{col_letra}{max_row}"
        ws.conditional_formatting.add(range_str, rule_enviado)
        ws.conditional_formatting.add(range_str, rule_atrasado)
        ws.conditional_formatting.add(range_str, rule_nao_possui)

    coluna_regional_idx = 1
    for row in range(2, ws.max_row + 1):
        cell_regional = ws.cell(row=row, column=coluna_regional_idx)
        nome_regional = cell_regional.value
        if nome_regional and nome_regional in cores_regionais:
            cor_hex = cores_regionais[nome_regional]
            fill_regional = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")
            cell_regional.fill = fill_regional

    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        if column_letter == 'C':
            continue
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column_letter].width = max_length + 5

    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.freeze_panes = 'D2'

    caminho_saida = pasta_outputs / f"{nome_convenio}_formB.xlsx"
    try:
        wb.save(caminho_saida)
        print(f"  - Planilha '{caminho_saida.name}' salva com sucesso!")
    except Exception as e:
        print(f"  - ERRO ao salvar '{caminho_saida.name}': {e}")


# --- LÓGICA PRINCIPAL DO SCRIPT ---

print("Iniciando o processo de geração de relatórios para o Formulário B...")

for convenio, caminho in caminhos_csv_formB.items():
    print(f"\nProcessando convênio: '{convenio.upper()}'")
    
    if not caminho.exists():
        print(f"  - AVISO: Arquivo '{caminho.name}' não encontrado. Pulando este convênio.")
        continue


    caminho_saida = pasta_outputs / f"{convenio}_formB.xlsx"
    df_estados_antigos = None
    
    # Verifica se a planilha de saída já existe para ler os dados dela
    if caminho_saida.exists():
        print(f"  - Encontrada planilha anterior. Lendo estados 'Não Possui'...")
        try:
            # Lê apenas as colunas chave e as de situação para otimizar
            colunas_para_ler = ['Município', 'UVR', 'Situação PS', 'Situação LR', 'Situação OS']
            df_estados_antigos = pd.read_excel(caminho_saida, usecols=colunas_para_ler, dtype=str)
            # Filtra para manter apenas as linhas que realmente foram marcadas como "Não Possui"
            df_estados_antigos = df_estados_antigos[
                (df_estados_antigos['Situação PS'] == 'Não Possui') |
                (df_estados_antigos['Situação LR'] == 'Não Possui') |
                (df_estados_antigos['Situação OS'] == 'Não Possui')
            ]
        except Exception as e:
            print(f"  - AVISO: Não foi possível ler a planilha anterior '{caminho_saida.name}': {e}")
            df_estados_antigos = None # Reseta em caso de erro

    df = pd.read_csv(caminho, dtype=str).fillna('---')
    
    df['municipio_chave'] = df['municipio'].apply(normalizar_texto)
    caminho_regional = caminhos_regionais_xlsx[convenio]
    
    if caminho_regional.exists():
        try:
            print(f"  - Carregando e padronizando dados de Regional de '{caminho_regional.name}'...")
            df_regional_lookup = pd.read_excel(
                caminho_regional, sheet_name="01.25",
                usecols=["Município", "UVR", "Regional"], dtype=str
            )
            df_regional_lookup.rename(columns={'Município': 'municipio', 'UVR': 'uvr'}, inplace=True)
            df_regional_lookup['municipio_chave'] = df_regional_lookup['municipio'].apply(normalizar_texto)
            df_regional_lookup.drop_duplicates(subset=["municipio_chave", "uvr"], inplace=True)
            
            df = pd.merge(
                df, df_regional_lookup[['municipio_chave', 'uvr', 'Regional']],
                on=['municipio_chave', 'uvr'], how='left'
            )
        except Exception as e:
            print(f"  - ERRO ao ler o arquivo de regionais '{caminho_regional.name}': {e}")
            df['Regional'] = 'ERRO NA LEITURA'
    else:
        print(f"  - AVISO: Arquivo de regionais '{caminho_regional.name}' não encontrado.")
        df['Regional'] = 'Não encontrado'

    df.drop(columns=['municipio_chave'], inplace=True, errors='ignore')
    df.fillna('---', inplace=True)

    mapa_colunas = {
        'municipio': 'Município', 'uvr': 'UVR',
        'regional_form_ps': 'Responsável PS', 'data_envio_form_ps': 'Data Envio PS',
        'regional_form_lr': 'Responsável LR', 'data_envio_form_lr': 'Data Envio LR',
        'regional_form_os': 'Responsável OS', 'data_envio_form_os': 'Data Envio OS'
    }
    df.rename(columns=mapa_colunas, inplace=True)
    
    servicos = [('PS', 'Data Envio PS'), ('LR', 'Data Envio LR'), ('OS', 'Data Envio OS')]
    
    for sigla, coluna_data in servicos:
        df[f'Situação {sigla}'] = df[coluna_data].apply(
            lambda x: 'Não Enviado' if str(x).strip() in ['---', ''] else 'Enviado'
        )
        df[coluna_data] = df[coluna_data].apply(formatar_data)

    ordem_final = [
        'Regional', 'Município', 'UVR',
        'Responsável PS', 'Situação PS', 'Data Envio PS',
        'Responsável LR', 'Situação LR', 'Data Envio LR',
        'Responsável OS', 'Situação OS', 'Data Envio OS'
    ]
    
    for col in ordem_final:
        if col not in df.columns:
            df[col] = '---'
            
    df_final = df[ordem_final]
    

    if df_estados_antigos is not None and not df_estados_antigos.empty:
        print("  - Mesclando dados novos com os estados 'Não Possui' salvos...")

        df_final.set_index(['Município', 'UVR'], inplace=True)
        df_estados_antigos.set_index(['Município', 'UVR'], inplace=True)
        
        df_final.update(df_estados_antigos)
        
        df_final.reset_index(inplace=True)

    gerar_planilha_estilizada_formB(df_final, convenio)

print("\nProcesso finalizado.")