from openpyxl import load_workbook, Workbook  # Para trabalhar com arquivos Excel
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment  # Para aplicar estilos nas células
from openpyxl.formatting.rule import CellIsRule
from copy import copy
from datetime import datetime  # Para manipular datas
import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path  # Para manipulação de caminhos de arquivos
from utils import (  # Estilos e funções auxiliares
    cabeçalho_fill, cabeçalho_font, enviado_fill, analise_fill, enviado_font,
    semtecnico_fill, atrasado_fill, validado_nao_fill, validado_sim_fill, atrasado2_fill, outras_fill, duplicado_fill,
    cores_regionais, bordas, alinhamento,
    normalizar_texto, aplicar_estilo_status
)

# Define o caminho do script atual
caminho_script = Path(__file__).resolve()
pasta_scripts = caminho_script.parent
pasta_inputs = pasta_scripts.parent / "inputs"

# Caminho do arquivo do banco e arquivos auxiliares (originais do drive)
csv_file_input = pasta_inputs/"form1.csv"
planilhas_auxiliares = {
    "belem": pasta_inputs / "0 - Belém" / "0 - Monitoramento Form 1, 2 e 3.xlsx",
    "expansao": pasta_inputs / "0 - Expansão" / "0 - Monitoramento Form 1, 2 e 3.xlsx",
    "grs": pasta_inputs / "0 - GRS II" / "0 - Monitoramento Form 1, 2 e 3.xlsx",
    "expansao_ms": pasta_inputs / "0 - Expansão MS" / "0 - Monitoramento Form 1, 2 e 3.xlsx"
}

# Carrega a planilha principal
df_input = pd.read_csv(csv_file_input, dtype=str)


# Dicionário para armazenar o status de envio por município
dados_atualizados = {}

for _, row in df_input.iterrows():
    municipio = row['municipio']
    data_envio = row['data_envio']


    # Normaliza o nome do município
    if isinstance(municipio, str):
        municipio_normalizado = normalizar_texto(municipio)
    else:
        continue

    # Tenta formatar a data de envio
    if isinstance(data_envio, datetime):
        data_envio_formatada = data_envio.strftime("%d/%m/%Y")
    else:
        try:
            data_envio_formatada = datetime.strptime(data_envio, "%Y-%m-%d %H:%M:%S.%f").strftime("%d/%m/%Y")
        except (ValueError, TypeError):
            data_envio_formatada = ""

    # Atualiza ou insere os dados no dicionário
    if municipio_normalizado in dados_atualizados:
        dados_atualizados[municipio_normalizado]["datas"].append(data_envio_formatada)
        dados_atualizados[municipio_normalizado]["status"] = "Duplicado"
    else:
        dados_atualizados[municipio_normalizado] = {"datas": [data_envio_formatada], "status": "Enviado"}

# Processa cada planilha auxiliar (belém, expansão, GRS)
for nome, caminho in planilhas_auxiliares.items():
    wb_aux = load_workbook(caminho)


    # Verifica se a aba existe
    if "Form 1 - Município" not in wb_aux.sheetnames:
        print(f"A aba 'Form 1 - Município' não foi encontrada em {nome}. Nenhuma modificação será feita.")
        continue

    ws_aux = wb_aux["Form 1 - Município"]

     
    # Usa o workbook que foi criado no EXECUTAR_TODOS
    wb_destino = {"belem": belem_wb, "expansao": expansao_wb, "grs": grs_wb, "expansao_ms": expansao_ms_wb}[nome] # type: ignore

    abas_para_copiar = ["Resumo", "Monitoramento"]
    for nome_aba in abas_para_copiar:
        if nome_aba in wb_aux.sheetnames:
            # Só copia se a aba ainda não existir no workbook de destino.
            if nome_aba not in wb_destino.sheetnames:
                print(f"Copiando a aba '{nome_aba}' do modelo '{nome}'...")
                ws_origem = wb_aux[nome_aba]
                ws_destino_aba = wb_destino.create_sheet(nome_aba)

                # Copia os dados e estilos célula por célula
                for row in ws_origem.iter_rows():
                    for cell in row:
                        new_cell = ws_destino_aba.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color)
                            new_cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom)
                            new_cell.fill = PatternFill(fill_type=cell.fill.fill_type, start_color=cell.fill.start_color, end_color=cell.fill.end_color)
                            new_cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=cell.alignment.wrap_text, shrink_to_fit=cell.alignment.shrink_to_fit)
                            new_cell.number_format = cell.number_format

                # Copia as dimensões das colunas e linhas
                for col_letter, dim in ws_origem.column_dimensions.items():
                    ws_destino_aba.column_dimensions[col_letter].width = dim.width
                for row_index, dim in ws_origem.row_dimensions.items():
                    ws_destino_aba.row_dimensions[row_index].height = dim.height

                # Copia as células mescladas
                for merged_cell_range in ws_origem.merged_cells.ranges:
                    ws_destino_aba.merge_cells(str(merged_cell_range))
                 
                # Copia as validações de dados
                for dv in ws_origem.data_validations.dataValidation:
                    ws_destino_aba.add_data_validation(dv)
                 
                # Copia a configuração do freeze_panes
                if ws_origem.freeze_panes:
                    ws_destino_aba.freeze_panes = ws_origem.freeze_panes


                for range_string in ws_origem.conditional_formatting:
                    rules_list = ws_origem.conditional_formatting[range_string]

                    for rule in rules_list:
                        ws_destino_aba.conditional_formatting.add(range_string, rule)




    novo_ws = wb_destino.create_sheet("Form 1 - Município")  

     

    dv_sim_nao = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True) #dropdown com sim e nao
    novo_ws.add_data_validation(dv_sim_nao)

    dv_sim_nao_ti = DataValidation(type="list", formula1='"Sim,Não, Em Análise"', allow_blank=True)
    novo_ws.add_data_validation(dv_sim_nao_ti)

    dv_status = DataValidation(type="list", formula1='"Enviado, Atrasado, Outras Ocorrências, Sem Técnico, Duplicado"', allow_blank=True) #dropdown de status
    novo_ws.add_data_validation(dv_status)

    # Copia e estiliza os cabeçalhos
    headers = [cell.value for cell in ws_aux[1]]
    for col_num, header in enumerate(headers, start=1):
        cell = novo_ws.cell(row=1, column=col_num, value=header)
        cell.fill = cabeçalho_fill
        cell.font = cabeçalho_font
        cell.border = bordas
        cell.alignment = alinhamento  

    novo_ws.auto_filter.ref = f"A1:G1"

    # Processa as linhas da planilha auxiliar

    for row_idx, row_cells in enumerate(ws_aux.iter_rows(min_row=2), start=2):

        municipio_original = row_cells[1].value
        row_data = [cell.value for cell in row_cells]

        # Normaliza o nome do município
        if isinstance(municipio_original, str):
            municipio_normalizado = normalizar_texto(municipio_original)
        else:
            municipio_normalizado = ""

        # Atualiza status e data de envio, se estiver no dicionário
        if municipio_normalizado in dados_atualizados:
             
            # 1. Contar o número de datas de envio na planilha anterior (coluna F, índice 5)
            datas_antigas_str = row_data[5] if row_data[5] and isinstance(row_data[5], str) else ""
            num_datas_antigas = len([data for data in datas_antigas_str.split(',') if data.strip()])

            # 2. Contar o número de novas datas de envio
            num_datas_novas = len(dados_atualizados[municipio_normalizado]["datas"])

            # 3. Se o número de datas aumentou, marcar "Validado pelo Regional" como "Não"
            if num_datas_novas > num_datas_antigas:
                row_data[6] = "Não"
             
            novas_datas = ", ".join(dados_atualizados[municipio_normalizado]["datas"])
            novo_status = dados_atualizados[municipio_normalizado]["status"]
            row_data[5] = novas_datas
            row_data[4] = novo_status
        else:
            # Caso não tenha envio:
            if row_data[4] == "Sem Técnico":
                pass
            elif row_data[4] is None:
                row_data[4] = "Atrasado"

        # Coloração de validação (Sim/Não)
        if row_data[6] == "Não":
            novo_ws.cell(row=row_idx, column=7).fill = validado_nao_fill
        elif row_data[6] == "Sim":
            novo_ws.cell(row=row_idx, column=7).fill = validado_sim_fill

        # Coloração regional
        regional = row_data[0]
        if regional in cores_regionais:
            cor_hex = cores_regionais[regional]
            novo_ws.cell(row=row_idx, column=1).fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")

        # Copia os dados para a nova planilha e aplica estilos
        for col_idx, value in enumerate(row_data, start=1):
            cell = novo_ws.cell(row=row_idx, column=col_idx, value=value)
            

            # Isso preserva quebras de linha (wrap_text) e outras formatações de alinhamento.
            original_cell = row_cells[col_idx - 1]

            if original_cell.has_style:
                cell.fill = copy(original_cell.fill)
            cell.alignment = Alignment(
                horizontal=original_cell.alignment.horizontal,
                vertical=original_cell.alignment.vertical,
                text_rotation=original_cell.alignment.text_rotation,
                wrap_text=original_cell.alignment.wrap_text,
                shrink_to_fit=original_cell.alignment.shrink_to_fit,
                indent=original_cell.alignment.indent
            )
            
            cell.border = bordas
            cell.font = Font(name='Arial', size=11)
                   
            if col_idx == 7:
                dv_sim_nao.add(cell.coordinate)

            if col_idx == 10:
                dv_sim_nao_ti.add(cell.coordinate)

            if col_idx == 5:
                dv_status.add(cell.coordinate) 


        # Coloração regional
        regional = row_data[0]
        if regional in cores_regionais:
            cor_hex = cores_regionais[regional]
            novo_ws.cell(row=row_idx, column=1).fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")       
                 

        # Isso garante que o conteúdo não fique escondido.
        source_row_index = row_cells[0].row
        if source_row_index in ws_aux.row_dimensions:
            novo_ws.row_dimensions[row_idx].height = ws_aux.row_dimensions[source_row_index].height

    # Aplica cor ao status 
    for row_idx in range(2, novo_ws.max_row + 1):
        status_cell = novo_ws.cell(row=row_idx, column=5)
        aplicar_estilo_status(status_cell, status_cell.value)

    # Ajusta a largura das colunas com base no conteúdo (pode ser ajustado ou removido se preferir manter as larguras originais)
    for col in novo_ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = col[0].column_letter
        # Para evitar que colunas com texto longo (como observações) fiquem excessivamente largas,
        # você pode definir uma largura máxima ou ajustar manualmente.
        # Por enquanto, a lógica original será mantida.
        novo_ws.column_dimensions[col_letter].width = max_length + 5


    novo_ws.freeze_panes = 'D1' #Congela as colunas A,B,C

        
    if novo_ws.max_row >= 2:
        coluna_validado_regional = f"G2:G{novo_ws.max_row}"
        coluna_validado_ti = f"J2:J{novo_ws.max_row}"
        coluna_status = f"E2:E{novo_ws.max_row}"
       
        rule_sim = CellIsRule(operator='equal', formula=['"Sim"'], stopIfTrue=True, fill=validado_sim_fill)
        novo_ws.conditional_formatting.add(coluna_validado_regional, rule_sim)
        novo_ws.conditional_formatting.add(coluna_validado_ti, rule_sim)

        rule_nao = CellIsRule(operator='equal', formula=['"Não"'], stopIfTrue=True, fill=validado_nao_fill)
        novo_ws.conditional_formatting.add(coluna_validado_regional, rule_nao)
        novo_ws.conditional_formatting.add(coluna_validado_ti, rule_nao)

        rule_analise = CellIsRule(operator='equal', formula=['"Em Análise"'], stopIfTrue=True, fill=analise_fill)
        novo_ws.conditional_formatting.add(coluna_validado_ti, rule_analise)


        status_rules = {
            "Enviado": {"fill": enviado_fill, "font": enviado_font},
            "Atrasado": {"fill": atrasado_fill, "font": enviado_font},
            "Outras Ocorrências": {"fill": outras_fill, "font": enviado_font},
            "Sem Técnico": {"fill": semtecnico_fill, "font": enviado_font},
            "Duplicado": {"fill": duplicado_fill, "font": enviado_font}
        }

        for status_text, styles in status_rules.items():
            rule = CellIsRule(operator='equal',
                            formula=[f'"{status_text}"'],
                            stopIfTrue=True,
                            fill=styles["fill"],
                            font=styles["font"])
            novo_ws.conditional_formatting.add(coluna_status, rule)        
     

    # Salva a nova planilha com nome específico
    #novo_caminho = pasta_scripts.parent / "form1" / f"{nome}_atualizado_form1.xlsx"
    #novo_wb.save(novo_caminho)