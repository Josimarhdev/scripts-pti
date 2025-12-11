import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
from openpyxl.styles import PatternFill
import unicodedata


try:
    from utils import (
        cabeçalho_fill, cabeçalho_font, enviado_fill, enviado_font,
        atrasado_fill, bordas, alinhamento, cores_regionais
    )
except ImportError:
    print("ERRO: O arquivo 'utils.py' não foi encontrado na pasta 'scripts'.")
    print("Por favor, certifique-se de que o seu arquivo utils.py está no lugar certo.")
    exit()


caminho_script = Path(__file__).resolve()
pasta_scripts = caminho_script.parent
pasta_inputs = pasta_scripts.parent / "inputs"
pasta_outputs = pasta_scripts.parent / "outputs"
pasta_outputs.mkdir(exist_ok=True)

caminhos_csv = {
    "belem": pasta_inputs / "formA-belém.csv",
    "expansao": pasta_inputs / "formA-expansão.csv",
    "grs": pasta_inputs / "formA-grs.csv"
}

caminhos_regionais_xlsx = {
    "belem": pasta_inputs / "0 - Belém" / "0 - Monitoramento Form 4.xlsx",
    "expansao": pasta_inputs / "0 - Expansão" / "0 - Monitoramento Form 4.xlsx",
    "grs": pasta_inputs / "0 - GRS II" / "0 - Monitoramento Form 4.xlsx"
}


def normalizar_texto(texto: str) -> str:
    if not isinstance(texto, str):
        return texto
    texto_sem_acento = ''.join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    )
    return texto_sem_acento.lower().strip()

def gerar_planilha_estilizada(df, nome_convenio):
    print(f"  - Criando Workbook para '{nome_convenio}'...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Monitoramento"
    
   
    colunas_finais = ['Regional', 'Município', 'UVR', 'Responsável', 'Situação', 'Data de Envio']
    ws.append(colunas_finais)

    # Aplica estilo ao cabeçalho
    for cell in ws[1]:
        cell.fill = cabeçalho_fill
        cell.font = cabeçalho_font
        cell.border = bordas
        cell.alignment = alinhamento

    # Adiciona dados e validação
    dv_status = DataValidation(type="list", formula1='"Enviado,Não Enviado"', allow_blank=True)
    ws.add_data_validation(dv_status)

    for index, row_data in df.iterrows():
        row_num = index + 2
        # Escreve os dados 
        for col_idx, col_name in enumerate(colunas_finais, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=row_data[col_name])
            cell.border = bordas
            cell.alignment = alinhamento
            if col_idx == 5:
                dv_status.add(cell)



    # Colore a coluna 'Situação' 
    coluna_situacao_idx = 5
    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row, column=coluna_situacao_idx)
        if status_cell.value == "Enviado":
            status_cell.fill = enviado_fill
            status_cell.font = enviado_font
        elif status_cell.value == "Não Enviado":
            status_cell.fill = atrasado_fill
            status_cell.font = enviado_font

    # Colore a coluna 'Regional'
    coluna_regional_idx = 1
    for row in range(2, ws.max_row + 1):
        cell_regional = ws.cell(row=row, column=coluna_regional_idx)
        nome_regional = cell_regional.value
        if nome_regional and nome_regional in cores_regionais:
            cor_hex = cores_regionais[nome_regional]
            fill_regional = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")
            cell_regional.fill = fill_regional

    # Ajusta a largura das colunas e congela painéis
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column_letter].width = max_length + 5


    ws.column_dimensions['B'].width = 35
    ws.freeze_panes = 'C2'

    # Salva o arquivo
    caminho_saida = pasta_outputs / f"{nome_convenio}_formA.xlsx"
    try:
        wb.save(caminho_saida)
        print(f"  - Planilha '{caminho_saida.name}' salva com sucesso!")
    except Exception as e:
        print(f"  - ERRO ao salvar '{caminho_saida.name}': {e}")


# --- LÓGICA PRINCIPAL DO SCRIPT ---
print("Iniciando o processo de geração de relatórios...")
for convenio, caminho_csv in caminhos_csv.items():
    print(f"\nProcessando convênio: '{convenio.upper()}'")
    
    if not caminho_csv.exists():
        print(f"  - AVISO: Arquivo CSV '{caminho_csv.name}' não encontrado. Pulando este convênio.")
        continue

    df = pd.read_csv(caminho_csv, dtype=str)
    df['municipio_chave'] = df['municipio'].apply(normalizar_texto)

    caminho_regional = caminhos_regionais_xlsx[convenio]
    if caminho_regional.exists():
        try:
            print(f"  - Carregando e padronizando dados de Regional de '{caminho_regional.name}'...")

  

   
            df_regional_lookup = pd.read_excel(
                caminho_regional, sheet_name="01.25",
                usecols=["Município", "UVR", "Regional"], 
                dtype=str
            )

         
            df_regional_lookup.rename(columns={'Município': 'municipio', 'UVR': 'uvr'}, inplace=True)

     
            df['uvr'] = pd.to_numeric(df['uvr'], errors='coerce')
            df_regional_lookup['uvr'] = pd.to_numeric(df_regional_lookup['uvr'], errors='coerce')

          
            df_regional_lookup['municipio_chave'] = df_regional_lookup['municipio'].apply(normalizar_texto)
            df_regional_lookup.drop_duplicates(subset=["municipio_chave", "uvr"], inplace=True)
            
            df = pd.merge(
                df, df_regional_lookup[['municipio_chave', 'uvr', 'Regional']],
                on=['municipio_chave', 'uvr'], how='left'
            )
  

        except Exception as e:
            print(f"  - ERRO ao ler o arquivo de regionais '{caminho_regional.name}': {e}")
            df['Regional'] = 'ERRO NA LEITURA'
    else:
        print(f"  - AVISO: Arquivo de regionais '{caminho_regional.name}' não encontrado.")
        df['Regional'] = 'Não encontrado'

    df.drop(columns=['municipio_chave'], inplace=True, errors='ignore')

    df.fillna('---', inplace=True)
    df.rename(columns={
        'regional': 'Responsável',
        'municipio': 'Município',
        'uvr': 'UVR',
        'data_envio': 'Data de Envio'
    }, inplace=True)

    df['Situação'] = df['Data de Envio'].apply(lambda x: 'Não Enviado' if str(x).strip() in ['---', ''] else 'Enviado')

    def formatar_data(data):
        if str(data).strip() in ['---', '']: return '---'
        try: return pd.to_datetime(data, dayfirst=True).strftime('%d/%m/%Y')
        except (ValueError, TypeError): return data
            
    df['Data de Envio'] = df['Data de Envio'].apply(formatar_data)
    
    colunas_finais_df = ['Regional', 'Município', 'UVR', 'Responsável', 'Situação', 'Data de Envio']
    df_final = df[colunas_finais_df]
    
    gerar_planilha_estilizada(df_final, convenio)

print("\nProcesso finalizado.")