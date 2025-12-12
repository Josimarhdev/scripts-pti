import pandas as pd
import re
import os
import sys # Necessário para ajustar o caminho dos imports
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import sys
import os
from utils import cores_regionais 



def processar_planilhas_excel():
    
    # Pega o diretório onde este script está salvo (.../engajamento)
    diretorio_do_script = os.path.dirname(os.path.abspath(__file__))
    caminho_dos_arquivos = os.path.join(diretorio_do_script, 'inputs')
    
    arquivo_geral = os.path.join(caminho_dos_arquivos, '0 - Monitoramento Form 1, 2 e 3.xlsx')
    arquivo_form4 = os.path.join(caminho_dos_arquivos, '0 - Monitoramento Form 4.xlsx')

    # Verificação para garantir que o script achou os arquivos
    if not os.path.exists(arquivo_geral) or not os.path.exists(arquivo_form4):
        print(f"Erro: Arquivos não encontrados na pasta:\n{caminho_dos_arquivos}")
        return None, None, None # Retorna 3 valores conforme esperado no main

    try:
        xls_geral = pd.ExcelFile(arquivo_geral)
        xls_form4 = pd.ExcelFile(arquivo_form4)
    except FileNotFoundError as e:
        print(f"Erro: Arquivo não encontrado.")
        print(f"Verifique se o caminho '{caminho_dos_arquivos}' e os nomes dos arquivos estão corretos.")
        return None, None, None 

    try:
        form1 = pd.read_excel(xls_geral, sheet_name='Form 1 - Município')
        form2 = pd.read_excel(xls_geral, sheet_name='Form 2 - UVR')
        form3 = pd.read_excel(xls_geral, sheet_name='Form 3 - Empreendimento')
        monitoramento = pd.read_excel(xls_geral, sheet_name='Monitoramento')
    except ValueError as e:
        print(f"Erro ao ler uma das abas: {e}. Verifique se os nomes das abas estão corretos.")
        return None, None, None 

    # --- PROCESSAMENTO DE DADOS ---
    monitoramento.drop(monitoramento[monitoramento['Regional'] == 'Regional'].index, inplace=True)
    monitoramento.rename(columns={'Municípios': 'Municipio'}, inplace=True)
    all_uvrs = monitoramento[['Regional', 'Municipio', 'UVR']].copy()
    all_uvrs.dropna(subset=['Regional', 'Municipio', 'UVR'], inplace=True)
    all_uvrs.drop_duplicates(inplace=True)
    all_uvrs.reset_index(drop=True, inplace=True)
    
    def check_submission(df, municipio, uvr):
        df_filtered = df[(df['Município'] == municipio) & (df['UVR'] == uvr)]
        if not df_filtered.empty:
            situacao = df_filtered['Situação'].iloc[0]
            if situacao in ['Enviado', 'Duplicado']:
                return 1
        return 0

    all_uvrs['Form 1'] = all_uvrs.apply(lambda row: check_submission(form1, row['Municipio'], row['UVR']), axis=1)
    all_uvrs['Form 2'] = all_uvrs.apply(lambda row: check_submission(form2, row['Municipio'], row['UVR']), axis=1)
    all_uvrs['Form 3'] = all_uvrs.apply(lambda row: check_submission(form3, row['Municipio'], row['UVR']), axis=1)
    
    current_date = datetime.now()
    current_year = current_date.year
    current_month = current_date.month

    expected_2024 = 2
    if current_year < 2025:
        expected_2025 = 0
    elif current_year == 2025:
        expected_2025 = current_month - 1
    else:
        expected_2025 = 12
    
    submissions_2024 = {}
    submissions_2025 = {}

    for sheet_name in xls_form4.sheet_names:
        if re.match(r'^\d{2}\.\d{2}$', sheet_name):
            try:
                df_month = pd.read_excel(xls_form4, sheet_name=sheet_name)
                year_suffix = sheet_name.split('.')[1]

                if year_suffix == '24' and sheet_name not in ['11.24', '12.24']:
                    continue
                
                if year_suffix == '24':
                    counts_dict = submissions_2024
                elif year_suffix == '25':
                    counts_dict = submissions_2025
                else:
                    continue
                
                for _, row in df_month.iterrows():
                    key = (row['Município'], row['UVR'])
                    if row['Situação'] in ['Enviado', 'Duplicado']:
                        counts_dict[key] = counts_dict.get(key, 0) + 1
            except Exception as e:
                print(f"Aviso: Não foi possível ler ou processar a aba '{sheet_name}'. Erro: {e}")

    all_uvrs['Form 4 2024'] = all_uvrs.apply(lambda row: submissions_2024.get((row['Municipio'], row['UVR']), 0), axis=1)
    all_uvrs['Form 4 2025'] = all_uvrs.apply(lambda row: submissions_2025.get((row['Municipio'], row['UVR']), 0), axis=1)

    total_expected = 3 + expected_2024 + expected_2025
    if total_expected > 0:
        all_uvrs['Total Submissions'] = all_uvrs['Form 1'] + all_uvrs['Form 2'] + all_uvrs['Form 3'] + all_uvrs['Form 4 2024'] + all_uvrs['Form 4 2025']
        all_uvrs['Engagement'] = (all_uvrs['Total Submissions'] / total_expected) * 100
    else:
        all_uvrs['Engagement'] = 0

    def get_engagement_level(percentage):
        if percentage > 90:
            return 'Alto'
        elif 60 <= percentage <= 90:
            return 'Médio'
        else:
            return 'Baixo'

    all_uvrs['Engagement Level'] = all_uvrs['Engagement'].apply(get_engagement_level)
    
    return all_uvrs, expected_2024, expected_2025


def criar_planilha_final(df, expected_2024, expected_2025): 
    """
    Cria a planilha Excel final com os dados de engajamento e formatação.
    """
    if df is None:
        print("Nenhum dado para processar. A planilha não foi criada.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Engajamento GRS"

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    greenS_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    yellowS_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    redS_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    white_fill = Font(color="FFFFFF", name='Arial', size=11)
    
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name='Arial', size=11)

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # --- Criação da header ---
    headers = [
        'Regional', 'Municipio', 'UVR', 'Form 1', 'Form 2', 'Form 3',
        f'Form 4 2024\n(Esperado: {expected_2024})',
        f'Form 4 2025\n(Esperado: {expected_2025})',
        'Nível de Engajamento'
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font 
        cell.border = thin_border
        cell.alignment = header_alignment 

    # --- Adição e Formatação dos Dados ---
    for _, row in df.iterrows():
        form1_status = 'Enviado' if row['Form 1'] == 1 else 'Ausente'
        form2_status = 'Enviado' if row['Form 2'] == 1 else 'Ausente'
        form3_status = 'Enviado' if row['Form 3'] == 1 else 'Ausente'
        
        data_row = [
            row['Regional'], row['Municipio'], row['UVR'],
            form1_status, form2_status, form3_status,
            row['Form 4 2024'], 
            row['Form 4 2025'], 
            row['Engagement Level']
        ]
        ws.append(data_row)
        
        current_row_index = ws.max_row
        
        for cell in ws[current_row_index]:
            cell.border = thin_border
            cell.alignment = center_alignment

        # Aplica a cor da Regional usando o dicionário importado
        regional_name = row['Regional']
        if regional_name in cores_regionais:
            color_code = cores_regionais[regional_name]
            regional_fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
            ws.cell(row=current_row_index, column=1).fill = regional_fill

        ws.cell(row=current_row_index, column=4).fill = green_fill if form1_status == 'Enviado' else red_fill
        ws.cell(row=current_row_index, column=5).fill = green_fill if form2_status == 'Enviado' else red_fill
        ws.cell(row=current_row_index, column=6).fill = green_fill if form3_status == 'Enviado' else red_fill

        level_cell = ws.cell(row=current_row_index, column=9)
        level = row['Engagement Level']
        if level == 'Alto':
            level_cell.fill = greenS_fill
            level_cell.font = white_fill
        elif level == 'Médio':
            level_cell.fill = yellowS_fill
            level_cell.font = white_fill
        else: #Baixo
            level_cell.fill = redS_fill
            level_cell.font = white_fill

    # --- Ajuste da Largura das Colunas ---
    ws.row_dimensions[1].height = 30 
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        
        if '\n' in str(col[0].value):
            max_length = max(len(part) for part in str(col[0].value).split('\n'))
        
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # --- AJUSTE DE SAÍDA DO ARQUIVO ---
    # Salva dentro da pasta 'engajamento/outputs'
    diretorio_atual = os.path.dirname(os.path.abspath(__file__))
    pasta_output = os.path.join(diretorio_atual, 'outputs')
    
    # Cria a pasta outputs se ela não existir (segurança)
    if not os.path.exists(pasta_output):
        os.makedirs(pasta_output)

    output_filename = os.path.join(pasta_output, "analise_engajamento.xlsx")
    
    wb.save(output_filename)
    print(f"\nPlanilha gerada com sucesso!")
    print(f"O arquivo foi salvo em: {os.path.abspath(output_filename)}")


# --- Execução Principal ---
if __name__ == "__main__":
    df_final, exp_2024, exp_2025 = processar_planilhas_excel()
    
    if df_final is not None:
        criar_planilha_final(
            df_final[[
                'Regional', 'Municipio', 'UVR', 'Form 1', 'Form 2', 'Form 3',
                'Form 4 2024', 'Form 4 2025', 'Engagement Level'
            ]],
            exp_2024,
            exp_2025
        )